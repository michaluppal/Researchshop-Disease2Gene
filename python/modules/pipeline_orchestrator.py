# modules/pipeline_orchestrator.py

import gzip
import json
import logging
import multiprocessing as mp
import os
import pickle
import time
import uuid

import pandas as pd
from tqdm import tqdm

from . import config, full_text_fetcher, pubmed_data_collector
from .abstract_screener import has_genetic_content
from .gemini_extractor import GeneInfoPipeline
from .pubtator_tool import HybridExtractionResult, NCBIGeneTool, PubTatorTool


class JobCancelledException(Exception):
    """Raised when a job is cancelled by the user."""

    pass


def _compute_row_confidence(row: dict, user_cols: list) -> tuple:
    """Compute a single confidence signal per output row.

    Returns (level, note) where level is HIGH/MEDIUM/LOW/REVIEW.
    - HIGH: gene confirmed by both NER+LLM sources, with verified citation
    - MEDIUM: passed all gates; citation stochastic or LLM-only
    - LOW: abstract-only paper or borderline validation confidence
    - REVIEW: citation mismatch detected OR gene from figure-only (no prose)
    """
    # Guard: empty gene name is never a valid extraction
    if not str(row.get("Gene/Group", "") or "").strip():
        return "REVIEW", "No genes extracted"

    val_conf = float(row.get("validation_confidence", 0) or 0)
    gene_source = str(row.get("Gene Source", "") or "")
    candidate_source = str(row.get("Candidate Source", "") or "")
    context_mods = str(row.get("context_modifications", "") or "")
    has_full_text = "no_oa_full_text" not in context_mods

    # Citation signals — check all user columns
    any_valid = False
    any_false = False
    all_empty = True
    for col in user_cols:
        citation_text = str(row.get(f"{col} Citation", "") or "")
        if citation_text:
            all_empty = False
        valid_flag = row.get(f"{col}_citation_valid")
        if valid_flag is True:
            any_valid = True
        elif valid_flag is False:
            any_false = True

    is_figure_only = (
        "llm_figure" in candidate_source
        and "llm_text" not in candidate_source
        and "pubtator" not in candidate_source
        and "deterministic_lexicon" not in candidate_source
    )

    # REVIEW: citation mismatch or figure-only source
    if (any_false and not any_valid) or is_figure_only:
        note = (
            "Figure-only gene — no prose citation available"
            if is_figure_only
            else "Citation text not found in paper"
        )
        return "REVIEW", note

    # LOW: no full text or borderline confidence
    if not has_full_text:
        return "LOW", "Abstract only"
    if val_conf < 0.85:
        return "LOW", "Low confidence"

    # HIGH: corroborated by multiple sources AND verified citation
    if gene_source == "both" and any_valid:
        return "HIGH", ""

    # MEDIUM: default (passed gates, citation stochastic or absent)
    if all_empty:
        note = ""
    elif not any_valid:
        note = "No citation"
    else:
        note = ""
    return "MEDIUM", note


logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - [%(module)s] - %(message)s"
)


def _run_pipeline_worker(text, cols, pubtator_genes=None, figure_inputs=None, abstract_text=None, table_inputs=None):
    """Top-level worker function for multiprocessing pool (must be picklable).

    Returns the result dict directly; mp.Pool.apply_async transmits it back to
    the orchestrator via IPC, eliminating the need for an explicit Queue.

    Args:
        text: Full paper text
        cols: User-defined column descriptions
        pubtator_genes: Optional list of gene symbols from PubTator (high-precision NER)
        figure_inputs: Optional list of figure metadata dicts for Gemini multimodal analysis
        abstract_text: Abstract text for independent abstract-level gene discovery
        table_inputs: Optional list of structured table dicts for table-cell citation validation
    """
    try:
        inst = GeneInfoPipeline(
            text,
            abstract_text=abstract_text or "",
            pubtator_genes=pubtator_genes,
            figure_inputs=figure_inputs,
            table_inputs=table_inputs,
        )
        df = inst.run_pipeline(cols)
        return {
            "records": df.to_dict(orient="records"),
            "debug": inst._collect_debug_artifact(),
            "gemini_api_calls": inst._paper_api_calls,
        }
    except Exception as e:
        import traceback

        logging.error(f"Pipeline worker error: {e}\n{traceback.format_exc()}")
        return {"error": str(e)}


def _create_unique_filepath(filename_base, extension):
    unique_id = uuid.uuid4().hex[:8]
    filename = f"{filename_base}_{unique_id}.{extension}"
    return os.path.join(config.OUTPUT_DIR, filename)


def _write_excel_output(
    df_clean: "pd.DataFrame",
    df_meta: "pd.DataFrame",
    excel_path: "os.PathLike",
) -> None:
    """Write a two-sheet Excel workbook: Results (clean) + Metadata (full).

    Sheet 1 'Results': clean researcher-facing columns with Confidence color coding.
    Sheet 2 'Metadata': full diagnostic/validation columns for auditing.
    Silently skips if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logging.warning(
            "openpyxl not installed — Excel output skipped. "
            "Install with: pip install openpyxl"
        )
        return

    _CONFIDENCE_FILLS = {
        "HIGH": "D4EDDA",
        "MEDIUM": "FFF9C4",
        "LOW": "FFE0B2",
        "REVIEW": "FCE4EC",
    }

    def _fill_sheet(ws, df: "pd.DataFrame", sheet_title: str, apply_conf_color: bool) -> None:
        ws.title = sheet_title
        headers = list(df.columns)
        conf_col_idx = (headers.index("Confidence") + 1) if (apply_conf_color and "Confidence" in headers) else None
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=False)
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                if conf_col_idx and col_idx == conf_col_idx:
                    hex_color = _CONFIDENCE_FILLS.get(str(value))
                    if hex_color:
                        cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)
        ws.freeze_panes = "A2"
        # Auto-size columns (sample first 100 rows for speed)
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            sample_rows = range(2, min(ws.max_row + 1, 102))
            max_len = max(
                len(str(header)),
                max((len(str(ws.cell(row=r, column=col_idx).value or "")) for r in sample_rows), default=0),
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb = Workbook()
    _fill_sheet(wb.active, df_clean, "Results", apply_conf_color=True)
    _fill_sheet(wb.create_sheet("Metadata"), df_meta, "Metadata", apply_conf_color=False)
    wb.save(excel_path)


def _write_json_output(df_clean: "pd.DataFrame", json_path: "os.PathLike") -> None:
    """Write the clean results as a JSON array of records."""
    df_clean.to_json(json_path, orient="records", indent=2, force_ascii=False)


def _write_split_output(
    df: "pd.DataFrame",
    output_path: "os.PathLike",
    user_cols: list,
) -> tuple:
    """Write primary CSV, metadata CSV, Excel workbook, and JSON file.

    Primary CSV: Gene, Variant, [user cols], Confidence, Confidence Note,
                 PMID, Title, Year, Journal, Authors, Citations, DOI
    Metadata CSV: PMID, Gene/Group, Variant Name + all diagnostic/validation columns
    Excel: two-sheet workbook (Results + Metadata), Confidence cells color-coded
    JSON: primary CSV data as an array of records

    Both CSV files share PMID + Gene/Group as join keys and have identical row order.

    Args:
        df: Full results DataFrame with all columns.
        output_path: Destination path for the primary CSV (.csv extension).
        user_cols: List of user-defined column names (without Citation pairs).

    Returns:
        (primary_path, metadata_path, excel_path, json_path) as strings.
    """
    from pathlib import Path

    output_path = Path(output_path)

    # Compute confidence flags row-by-row
    confidence_rows = [
        _compute_row_confidence(row, user_cols)
        for row in df.to_dict("records")
    ]
    df = df.copy()
    df["Confidence"] = [r[0] for r in confidence_rows]
    df["Confidence Note"] = [r[1] for r in confidence_rows]

    # --- Primary CSV (clean researcher-facing) ---
    rename_map = {
        "Gene/Group": "Gene",
        "Variant Name": "Variant",
        "Study Title": "Title",
        "Publication Year": "Year",
        "Journal Name": "Journal",
    }
    df_clean = df.rename(columns=rename_map)

    primary_cols = (
        ["Gene", "Variant"]
        + [c for c in user_cols if c in df_clean.columns]
        + ["Confidence", "Confidence Note", "context_modifications"]  # context_modifications included so
        # researchers working with the primary CSV can see what sections were truncated per gene row
        + ["PMID", "Title", "Year", "Journal", "Authors", "Citations", "DOI"]
    )
    primary_cols_present = [c for c in primary_cols if c in df_clean.columns]
    df_primary = df_clean[primary_cols_present]
    df_primary.to_csv(output_path, index=False)

    # --- Metadata CSV (full transparency, same row order as primary) ---
    meta_path = output_path.with_name(output_path.stem + "_metadata.csv")
    join_keys = ["PMID", "Gene/Group", "Variant Name"]
    remaining = [c for c in df.columns if c not in join_keys]
    meta_cols_present = [c for c in join_keys + remaining if c in df.columns]
    df_meta = df[meta_cols_present]
    df_meta.to_csv(meta_path, index=False)

    # --- Excel workbook (Results + Metadata sheets) ---
    excel_path = output_path.with_suffix(".xlsx")
    _write_excel_output(df_primary, df_meta, excel_path)

    # --- JSON (primary data as records) ---
    json_path = output_path.with_suffix(".json")
    _write_json_output(df_primary, json_path)

    return str(output_path), str(meta_path), str(excel_path), str(json_path)


def _sanitize_user_columns(user_columns: dict) -> dict:
    """Rename user-provided columns that collide with reserved/core names.

    Ensures the keys sent to the extractor are unique and do not overlap with
    core columns like 'Gene/Group', 'PMID', etc. If a collision is detected,
    the column is renamed to 'User: <name>'.
    """
    if not user_columns:
        return {}

    reserved_core = {
        "Gene/Group",
        "Variant Name",
        "PMID",
        "Study Title",
        "Authors",
        "Publication Year",
        "Journal Name",
        "Author Affiliations",
        "Citations",
        "Abstract",
        # Internal identifiers occasionally surfaced
        "gene_name",
        "variant_name",
    }

    out: dict = {}
    used: set = set()
    for raw_key, desc in user_columns.items():
        key = (raw_key or "").strip()
        if not key:
            continue
        # Avoid direct collisions with reserved core or already-used keys
        candidate = key
        if candidate in reserved_core or candidate in used:
            candidate = f"User: {candidate}"
        # If still collides, append numeric suffix until unique
        suffix = 2
        while candidate in reserved_core or candidate in used:
            candidate = f"{key} ({suffix})"
            suffix += 1
        out[candidate] = desc
        used.add(candidate)
    return out


def _ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure DataFrame has uniquely named columns by suffixing duplicates.

    Pandas reindexing requires unique column names; this function disambiguates
    duplicates by appending " (2)", "(3)", ... in appearance order.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    cols = list(df.columns)
    seen_counts: dict[str, int] = {}
    new_cols: list[str] = []
    for c in cols:
        count = seen_counts.get(c, 0)
        if count == 0:
            new_cols.append(c)
        else:
            new_cols.append(f"{c} ({count + 1})")
        seen_counts[c] = count + 1
    df.columns = new_cols
    return df


def run_complete_pipeline(
    query,
    specific_pmids,
    specific_authors,
    user_columns,
    top_n_cited,
    progress_callback=None,
    log_callback=None,
):
    """
    Runs the complete pipeline with the provided parameters.

    Args:
        log_callback: Optional callback(level, msg, detail) for structured log events.
                      Levels: "info" (user-facing), "debug" (technical), "warn", "error"
    """
    logging.info("--- STARTING COMPLETE PIPELINE RUN ---")
    start_time = time.time()

    def emit_log(level, msg, detail=None):
        """Emit a structured log event to the frontend."""
        if log_callback:
            log_callback(level, msg, detail)
        # Also route to Python logging for stderr capture
        log_level = {
            "info": logging.INFO,
            "debug": logging.DEBUG,
            "warn": logging.WARNING,
            "error": logging.ERROR,
        }.get(level, logging.INFO)
        logging.log(log_level, msg)

    def check_cancellation():
        """Check if job was cancelled via local .cancel file."""
        cancel_path = os.path.join(config.OUTPUT_DIR, ".cancel")
        if os.path.exists(cancel_path):
            emit_log("warn", "Pipeline cancelled by user")
            raise JobCancelledException("Job was cancelled")

    # Track pipeline statistics for frontend display
    pipeline_stats = {
        "papers_found": 0,
        "papers_screened": 0,
        "papers_screened_passed": 0,
        "papers_fetch_failed": 0,
        "papers_analyzed": 0,
        "genes_extracted": 0,
        "gemini_api_calls": 0,
        "tables_extracted": 0,
        "pubtator_pmids_skipped": 0,
    }
    paper_debug_artifacts = []
    forensic_screening = []
    fetch_report = []

    def get_citation_record(pmid, citation_records):
        rec = citation_records.get(pmid, {}) if citation_records else {}
        return {
            "count": int(rec.get("count", 0) or 0),
            "source": rec.get("source", "none") or "none",
            "retrieved_at": rec.get("retrieved_at", "") or "",
            "icite_count": rec.get("icite_count"),
            "semantic_scholar_count": rec.get("semantic_scholar_count"),
        }

    def build_minimal_row(pmid, base_info, citation_records, gene_group="", variant_name=""):
        citation = get_citation_record(pmid, citation_records)
        return {
            "Gene/Group": gene_group,
            "Variant Name": variant_name,
            "Candidate Source": "",
            "Normalization Applied": "",
            "Validation Outcome": "",
            "Dropped By Gate": "",
            "PMID": pmid,
            "DOI": base_info.get("doi", ""),
            "Study Title": base_info.get("title", "N/A"),
            "Authors": ", ".join(base_info.get("authors", [])),
            "Publication Year": base_info.get("year", "N/A"),
            "Journal Name": base_info.get("journal", "N/A"),
            "Author Affiliations": "; ".join(base_info.get("affiliations", [])),
            "Citations": citation["count"],
            "Citation Source": citation["source"],
            "Citation Retrieved At": citation["retrieved_at"],
            "iCite Citations": citation["icite_count"],
            "Semantic Scholar Citations": citation["semantic_scholar_count"],
            "Figure Count": 0,
            "Figure Analysis Enabled": bool(getattr(config, "ENABLE_FIGURE_ANALYSIS", True)),
            "Abstract": base_info.get("abstract", "No abstract available"),
            "Metadata Completeness": base_info.get("_metadata_completeness", 0),
            "Metadata Warnings": "; ".join(base_info.get("_metadata_warnings", [])),
        }

    def report_progress(stage, pct, extra_stats=None):
        check_cancellation()
        # Merge extra stats into pipeline_stats
        if extra_stats:
            pipeline_stats.update(extra_stats)
        if progress_callback:
            progress_callback(stage, pct, pipeline_stats.copy())
        logging.info(f"Progress: {stage} ({pct}%) - Stats: {pipeline_stats}")

    def write_drop_debug_artifact(status: str, output_csv_path: str = ""):
        """
        Persist per-run candidate/drop diagnostics to JSON for trust debugging.

        Includes forensic data when available: screening decisions, fetch outcomes,
        and table extraction statistics.
        """
        payload = {
            "status": status,
            "generated_at_epoch": time.time(),
            "query": query,
            "specific_pmids": list(specific_pmids or []),
            "specific_authors": list(specific_authors or []),
            "top_n_cited": top_n_cited,
            "output_csv_path": output_csv_path or "",
            "paper_debug": paper_debug_artifacts,
            "pipeline_stats": pipeline_stats.copy(),
            "screening_decisions": forensic_screening,
            "fetch_outcomes": fetch_report,
        }
        debug_path = _create_unique_filepath("drop_debug", "json")
        try:
            with open(debug_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
            emit_log("info", "Saved drop-debug artifact", debug_path)
            return debug_path
        except Exception as e:
            logging.warning(f"Failed to write drop-debug artifact: {e}")
            return ""

    report_progress("Initializing pipeline", 0)
    emit_log("info", "Initializing pipeline")

    # Step 1 & 2: Get PMIDs and Details
    report_progress("Searching PubMed", 10)
    if query:
        emit_log("info", f"Searching PubMed for: {query}")
    if specific_pmids:
        emit_log("info", f"Including {len(specific_pmids)} specific PMIDs")
    if specific_authors:
        emit_log("info", f"Searching papers by {len(specific_authors)} author(s)")
    logging.info("STEP 1 & 2: Gathering and fetching details for all PMIDs...")

    initial_pmids = set()
    mandatory_pmids = set(specific_pmids)  # Start with explicit PMIDs
    author_pmids = set()

    if specific_authors:
        for author in specific_authors:
            check_cancellation()
            author_results = pubmed_data_collector.search_pubmed_by_author(author, max_results=200)
            initial_pmids.update(author_results)
            author_pmids.update(author_results)

    mandatory_pmids |= author_pmids  # Union all author-derived PMIDs as mandatory

    # Always include explicitly requested PMIDs
    if mandatory_pmids:
        initial_pmids.update(mandatory_pmids)

    if query:
        check_cancellation()
        if getattr(config, "ENABLE_OA_FILTER", True):
            emit_log(
                "info",
                "Filtering to open-access papers only \u2014 paywalled papers will be excluded",
                "Filter: loattrfull text[sb]",
            )
        # Strategy: get the 100 most relevant (configurable), then later rank those by citations
        relevant_count = getattr(config, "PUBMED_RELEVANT_COUNT", 100)
        query_results = pubmed_data_collector.search_pubmed(query, relevant_count)
        initial_pmids.update(query_results)

    emit_log("info", f"Found {len(initial_pmids)} papers")
    report_progress("Fetching paper details", 20, {"papers_found": len(initial_pmids)})
    check_cancellation()
    emit_log("debug", f"Fetching details for {len(initial_pmids)} PMIDs from PubMed")
    paper_details = pubmed_data_collector.fetch_paper_details(list(initial_pmids))
    incomplete_count = sum(
        1
        for v in paper_details.values()
        if (v.get("_metadata_completeness", 1) < 1 or v.get("_metadata_warnings"))
    )
    if incomplete_count:
        emit_log(
            "warn",
            f"Metadata quality warning: {incomplete_count}/{len(paper_details)} papers have incomplete metadata",
            "Check 'Metadata Completeness' and 'Metadata Warnings' columns in output",
        )

    # Step 3: Fetch Full Text first for relevance-selected PMIDs
    report_progress("Fetching full text", 30)
    emit_log("info", f"Fetching full text for {len(paper_details)} papers")
    logging.info("STEP 3: Fetching full text for relevance-selected PMIDs...")
    check_cancellation()
    pmids_to_fetch = list(paper_details.keys())
    content_dict_path = _create_unique_filepath("content_dict", "pkl.gz")
    full_text_fetcher.run_fetching(pmids_to_fetch, content_dict_path)

    report_progress("Processing fetched content", 45)
    try:
        with gzip.open(content_dict_path, "rb") as f:
            content_dict = pickle.load(f)
    except Exception:
        logging.warning("Content dictionary is empty. No papers could be fetched.")
        return None

    # Keep only PMIDs successfully scraped
    scraped_pmids = list(content_dict.keys())
    emit_log(
        "info", f"Retrieved full text for {len(scraped_pmids)} of {len(pmids_to_fetch)} papers"
    )
    if getattr(config, "ENABLE_FIGURE_ANALYSIS", True):
        papers_with_figures = 0
        total_figures = 0
        for payload in content_dict.values():
            figs = payload.get("figures", []) if isinstance(payload, dict) else []
            if figs:
                papers_with_figures += 1
                total_figures += len(figs)
        if total_figures:
            emit_log("info", f"Discovered {total_figures} figures in {papers_with_figures} papers")

    # Count tables extracted from content_dict
    total_tables = sum(
        len(v.get("tables", []))
        for v in content_dict.values()
        if isinstance(v, dict)
    )
    pipeline_stats["tables_extracted"] = total_tables
    if total_tables:
        papers_with_tables = sum(
            1
            for v in content_dict.values()
            if isinstance(v, dict) and v.get("tables")
        )
        emit_log("info", f"Extracted {total_tables} tables from {papers_with_tables} papers")

    # Track fetch failures
    pipeline_stats["papers_fetch_failed"] = len(pmids_to_fetch) - len(scraped_pmids)

    # Forensic fetch report
    if getattr(config, "FORENSIC_INCLUDE_FETCH_OUTCOMES", True) and content_dict:
        try:
            from .full_text_fetcher import generate_fetch_report

            fetch_report.extend(generate_fetch_report(content_dict))
        except Exception as e:
            logging.warning(f"Forensic fetch report generation failed: {e}")

    # HYBRID PIPELINE: PubTator extraction for high-precision NER
    pubtator_results = {}  # pmid -> HybridExtractionResult
    if getattr(config, "ENABLE_PUBTATOR_EXTRACTION", True) and scraped_pmids:
        report_progress("Running PubTator extraction", 47)
        emit_log("info", "Running PubTator gene extraction")
        logging.info("STEP 3.5: Running PubTator extraction for high-precision gene discovery...")
        check_cancellation()

        try:
            pubtator_tool = PubTatorTool()
            batch_results = pubtator_tool.extract_from_pmids(scraped_pmids)

            for pmid, (genes, variants) in batch_results.items():
                result = HybridExtractionResult(pmid)
                result.pubtator_genes = genes
                result.pubtator_variants = variants
                pubtator_results[pmid] = result

            total_pt_genes = sum(len(r.pubtator_genes) for r in pubtator_results.values())
            pt_skipped = len(scraped_pmids) - len(pubtator_results)
            pipeline_stats["pubtator_pmids_skipped"] = pt_skipped
            emit_log(
                "info",
                f"PubTator found {total_pt_genes} genes across {len(pubtator_results)} papers",
                f"{pt_skipped} PMID(s) not returned by PubTator (not indexed or parse error)" if pt_skipped else None,
            )
            logging.info(
                f"PubTator: Found {total_pt_genes} genes across {len(pubtator_results)} papers"
            )
        except Exception as e:
            emit_log("warn", "PubTator extraction failed, continuing without it")
            logging.warning(f"PubTator extraction failed, continuing without it: {e}")

    if not scraped_pmids:
        emit_log("warn", "No full text available, falling back to metadata-only results")
        logging.warning(
            "No scraped papers available after full-text step. Falling back to minimal rows from PubMed metadata."
        )
        # Build minimal rows for initial PMIDs using PubMed details and citation counts
        try:
            # Fetch citations for all initial candidates
            all_candidate_pmids = list(paper_details.keys())
            citation_records = pubmed_data_collector.fetch_citation_counts_with_fallback(
                all_candidate_pmids
            )
            citations_dict = {pmid: rec.get("count", 0) for pmid, rec in citation_records.items()}

            # Rank by citations desc and take top_n_cited
            ranked = sorted(
                all_candidate_pmids, key=lambda p: citations_dict.get(p, 0), reverse=True
            )
            selected = ranked[:top_n_cited] if top_n_cited else ranked

            minimal_rows_only = []
            for pmid in selected:
                base_info = paper_details.get(pmid, {})
                minimal_rows_only.append(build_minimal_row(pmid, base_info, citation_records))

            all_results_df = pd.DataFrame(minimal_rows_only)

            # Save immediately (no user columns beyond core ones yet)
            output_path = _create_unique_filepath("final_enriched_results", "csv")
            all_results_df.to_csv(output_path, index=False, encoding="utf-8")
            debug_path = write_drop_debug_artifact(
                status="metadata_only_no_full_text", output_csv_path=output_path
            )
            report_progress("Completed", 100)
            logging.info(
                f"--- PIPELINE FINISHED WITH MINIMAL ROWS IN {time.time() - start_time:.2f} SECONDS ---"
            )
            return {"local_path": output_path, "debug_path": debug_path}
        except Exception as e:
            logging.error(f"Fallback to minimal rows failed: {e}")
            return None

    # Step 4: Fetch Citations only for scraped PMIDs and select top for AI
    report_progress("Fetching citation counts", 50)
    logging.info("STEP 4: Fetching citations for scraped PMIDs and selecting top papers...")
    check_cancellation()
    citation_records = pubmed_data_collector.fetch_citation_counts_with_fallback(scraped_pmids)
    citations_dict = {pmid: rec.get("count", 0) for pmid, rec in citation_records.items()}

    scraped_details = {pmid: info for pmid, info in paper_details.items() if pmid in scraped_pmids}
    all_papers_df = pd.DataFrame.from_dict(scraped_details, orient="index")
    if "citations" not in all_papers_df.columns:
        all_papers_df["citations"] = 0
    for pmid, count in citations_dict.items():
        if pmid in all_papers_df.index:
            all_papers_df.at[pmid, "citations"] = count

    report_progress("Selecting top papers", 60)
    # Build processing order: mandatory (if scraped) keep first (original order), then remaining by citations desc
    scraped_set = set(all_papers_df.index)
    ordered_mandatory = [p for p in list(mandatory_pmids) if p in scraped_set]
    remaining_df = (
        all_papers_df.drop(ordered_mandatory, errors="ignore")
        if ordered_mandatory
        else all_papers_df
    )
    ranked_remaining = remaining_df.sort_values(by="citations", ascending=False).index.tolist()
    pmids_to_process = ordered_mandatory + ranked_remaining
    logging.info(
        f"Prepared {len(pmids_to_process)} scraped PMIDs for AI analysis (mandatory first, then ranked by citations)."
    )

    # Abstract screening — forensic logging only (filtering moved to UI preview).
    # The UI now scores papers before the user selects them, so all user-selected
    # papers proceed to AI analysis without silent filtering.
    if getattr(config, "ENABLE_ABSTRACT_SCREENING", True):
        report_progress("Scoring abstracts", 65)
        logging.info(
            "STEP 4.5: Scoring paper abstracts for forensic logging (no filtering — screening moved to UI)..."
        )
        check_cancellation()

        threshold = getattr(config, "ABSTRACT_SCREENING_THRESHOLD", 5)
        would_pass = 0
        would_reject = 0

        for pmid in pmids_to_process:
            base_info = paper_details.get(pmid, {})
            title = base_info.get("title", "")
            abstract = base_info.get("abstract", "")
            should_process, confidence, details = has_genetic_content(abstract, title, threshold)
            if should_process or pmid in mandatory_pmids:
                would_pass += 1
            else:
                would_reject += 1

        logging.info(
            f"Abstract scoring: {would_pass}/{len(pmids_to_process)} would pass threshold "
            f"(all {len(pmids_to_process)} papers proceeding — screening moved to UI)"
        )
        # All papers proceed — no filtering
        pipeline_stats["papers_screened_passed"] = len(pmids_to_process)
        report_progress("Scoring abstracts", 68, {"papers_screened": len(pmids_to_process)})

        # Forensic screening data collection (still useful for debug artifacts)
        if getattr(config, "FORENSIC_INCLUDE_SCREENING", True):
            try:
                from .abstract_screener import screen_papers_with_decisions, decisions_to_dicts

                screening_subset = {
                    pmid: paper_details.get(pmid, {}) for pmid in pmids_to_process
                }
                _, screening_decisions = screen_papers_with_decisions(
                    screening_subset,
                    threshold=threshold,
                    mandatory_pmids=mandatory_pmids,
                )
                forensic_screening.extend(decisions_to_dicts(screening_decisions))
            except Exception as e:
                logging.warning(f"Forensic screening data collection failed: {e}")

        if not pmids_to_process:
            logging.warning(
                "No papers to analyse. Falling back to minimal rows."
            )
            # Build minimal rows for all scraped PMIDs
            minimal_rows_only = []
            for pmid in scraped_pmids[:top_n_cited]:
                base_info = paper_details.get(pmid, {})
                minimal_rows_only.append({**build_minimal_row(pmid, base_info, citation_records)})
            all_results_df = pd.DataFrame(minimal_rows_only)
            output_path = _create_unique_filepath("final_enriched_results", "csv")
            all_results_df.to_csv(output_path, index=False, encoding="utf-8")
            debug_path = write_drop_debug_artifact(
                status="metadata_only_all_rejected_by_screening", output_csv_path=output_path
            )
            report_progress("Completed", 100)
            logging.info(
                f"--- PIPELINE FINISHED WITH MINIMAL ROWS IN {time.time() - start_time:.2f} SECONDS ---"
            )
            return {"local_path": output_path, "debug_path": debug_path}

    report_progress("Analyzing papers with AI", 70)
    emit_log("info", f"Analyzing {len(pmids_to_process)} papers with AI")
    # Step 5: Process each paper using the GeneInfoPipeline class (Gemini)
    logging.info("STEP 5: Analyzing papers with Gemini using the GeneInfoPipeline...")
    all_results_df = pd.DataFrame()
    # Sanitize user columns to avoid collisions with core fields
    column_descriptions = _sanitize_user_columns(user_columns)

    total_papers = len(pmids_to_process)
    collected_rows = []
    full_rows_pmids = set()
    minimal_rows = []
    analyzed_attempts = 0

    # Pre-warm persistent worker pool — processes stay alive between papers so the
    # Python interpreter + heavy imports (google-genai, grpc, pandas) are paid once,
    # not once per paper. pool_size=2 means a timed-out worker doesn't block the next paper.
    pool_size = max(1, min(int(getattr(config, "AI_WORKER_POOL_SIZE", 2)), 4))
    worker_pool = mp.Pool(processes=pool_size)
    logging.info(f"AI worker pool created: {pool_size} processes")

    try:
        for i, pmid in enumerate(tqdm(pmids_to_process, desc="Processing papers")):
            analyzed_attempts += 1
            # Report progress for AI analysis phase (70-95%)
            ai_progress = 70 + int((i / total_papers) * 25)
            # This call raises JobCancelledException if cancellation signal detected
            report_progress(
                "Analyzing papers with AI", ai_progress, {"papers_analyzed": analyzed_attempts}
            )

            # Emit per-paper log
            base_title = paper_details.get(pmid, {}).get("title", "")
            short_title = (base_title[:60] + "...") if len(base_title) > 60 else base_title
            emit_log(
                "info", f"Analyzing paper {i + 1}/{total_papers}: {short_title}", f"PMID {pmid}"
            )

            content = content_dict.get(pmid, {})
            paper_text = content.get("content", "")
            if not paper_text:
                # No extracted content: create a minimal row so the study is still represented
                base_info = paper_details.get(pmid, {})
                minimal_rows.append(build_minimal_row(pmid, base_info, citation_records))
                paper_debug_artifacts.append(
                    {
                        "pmid": pmid,
                        "status": "no_full_text",
                        "reason": "missing_paper_text_after_fetch",
                        "emitted_rows": 0,
                    }
                )
                continue

            figure_inputs = (
                content.get("figures", [])
                if getattr(config, "ENABLE_FIGURE_ANALYSIS", True)
                else []
            )

            table_inputs = (
                content.get("tables", [])
                if getattr(config, "ENABLE_TABLE_CITATIONS", True)
                else []
            )

            base_info = paper_details.get(pmid, {})
            abstract = base_info.get("abstract", "")
            title = base_info.get("title", "")

            # Get PubTator genes for this paper (if available) to pass to LLM
            pt_gene_symbols = []
            if pmid in pubtator_results:
                pt_gene_symbols = [g.symbol for g in pubtator_results[pmid].pubtator_genes]
                if pt_gene_symbols:
                    logging.debug(
                        f"PMID {pmid}: Passing {len(pt_gene_symbols)} PubTator genes to LLM"
                    )

            paper_df = pd.DataFrame()
            worker_debug = {}
            try:
                ar = worker_pool.apply_async(
                    _run_pipeline_worker,
                    args=(
                        paper_text,
                        column_descriptions,
                        pt_gene_symbols,
                        figure_inputs,
                        abstract,
                        table_inputs,
                    ),
                )
                try:
                    payload = ar.get(timeout=config.AI_PER_PAPER_TIMEOUT_SECONDS)
                except mp.TimeoutError:
                    emit_log("warn", f"AI analysis timed out for PMID {pmid}, skipping")
                    logging.warning(
                        f"AI analysis timed out for PMID {pmid} after {config.AI_PER_PAPER_TIMEOUT_SECONDS}s; skipping"
                    )
                    worker_debug = {
                        "status": "timeout",
                        "reason": f"ai_timeout_{config.AI_PER_PAPER_TIMEOUT_SECONDS}s",
                    }
                    # Terminate the stuck worker(s) and create a fresh pool so subsequent
                    # papers aren't blocked waiting for an unresponsive worker process.
                    try:
                        worker_pool.terminate()
                        worker_pool.join(timeout=10)
                    except Exception as e:
                        logging.warning(f"Worker pool cleanup after timeout failed: {e}")
                    worker_pool = mp.Pool(processes=pool_size)
                    payload = None
                if payload is not None:
                    if isinstance(payload, dict) and payload.get("error"):
                        logging.error(f"AI analysis error for PMID {pmid}: {payload['error']}")
                        worker_debug = {
                            "status": "worker_error",
                            "reason": str(payload.get("error")),
                        }
                        paper_df = pd.DataFrame()
                    elif isinstance(payload, dict) and payload.get("records") is not None:
                        paper_df = pd.DataFrame(payload["records"])  # reconstruct DataFrame
                        if isinstance(payload.get("debug"), dict):
                            worker_debug = payload["debug"]
                        else:
                            worker_debug = {"status": "ok"}
                        pipeline_stats["gemini_api_calls"] += int(
                            payload.get("gemini_api_calls", 0)
                        )
                        # Surface context window warnings to the user
                        ctx_warn = (worker_debug or {}).get("context_warning")
                        if ctx_warn:
                            emit_log("warn", f"PMID {pmid}: {ctx_warn}")
            except Exception as e:
                emit_log("error", f"AI analysis failed for PMID {pmid}", str(e))
                logging.error(f"Failed AI analysis for PMID {pmid}: {e}")
                worker_debug = {
                    "status": "orchestrator_error",
                    "reason": str(e),
                }
                paper_df = pd.DataFrame()

            # Rename columns to match user requirements
            paper_df = paper_df.rename(
                columns={"gene_name": "Gene/Group", "variant_name": "Variant Name"}
            )

            base_info = paper_details.get(pmid, {})
            paper_df["PMID"] = pmid
            paper_df["DOI"] = base_info.get("doi", "")
            paper_df["Study Title"] = base_info.get("title", "N/A")
            paper_df["Authors"] = ", ".join(base_info.get("authors", []))
            paper_df["Publication Year"] = base_info.get("year", "N/A")
            paper_df["Journal Name"] = base_info.get("journal", "N/A")
            paper_df["Author Affiliations"] = "; ".join(base_info.get("affiliations", []))
            citation = get_citation_record(pmid, citation_records)
            paper_df["Citations"] = citation["count"]
            paper_df["Citation Source"] = citation["source"]
            paper_df["Citation Retrieved At"] = citation["retrieved_at"]
            paper_df["iCite Citations"] = citation["icite_count"]
            paper_df["Semantic Scholar Citations"] = citation["semantic_scholar_count"]
            paper_df["Metadata Completeness"] = base_info.get("_metadata_completeness", 0)
            paper_df["Metadata Warnings"] = "; ".join(base_info.get("_metadata_warnings", []))
            paper_df["Figure Count"] = len(figure_inputs)
            paper_df["Figure Analysis Enabled"] = bool(
                getattr(config, "ENABLE_FIGURE_ANALYSIS", True)
            )

            # HYBRID PIPELINE: Add gene source tracking
            if not paper_df.empty and pmid in pubtator_results:
                hybrid_result = pubtator_results[pmid]
                pt_symbols = {g.symbol.upper() for g in hybrid_result.pubtator_genes}

                # Track LLM genes in hybrid result
                if "Gene/Group" in paper_df.columns:
                    llm_genes = paper_df["Gene/Group"].dropna().unique().tolist()
                    hybrid_result.llm_genes = [g for g in llm_genes if g]

                # Add gene source column
                def get_gene_source(gene):
                    if not gene:
                        return ""
                    gene_upper = str(gene).upper()
                    in_pubtator = gene_upper in pt_symbols
                    in_llm = gene_upper in {str(g).upper() for g in hybrid_result.llm_genes}
                    if in_pubtator and in_llm:
                        return "both"
                    elif in_pubtator:
                        return "pubtator"
                    else:
                        return "llm"

                paper_df["Gene Source"] = paper_df["Gene/Group"].apply(get_gene_source)

                # Add NCBI Gene ID from PubTator
                def get_ncbi_id(gene):
                    if not gene:
                        return ""
                    gene_upper = str(gene).upper()
                    for g in hybrid_result.pubtator_genes:
                        if g.symbol.upper() == gene_upper and g.ncbi_gene_id:
                            return g.ncbi_gene_id
                    return ""

                paper_df["NCBI Gene ID"] = paper_df["Gene/Group"].apply(get_ncbi_id)

            # Use actual paper data for basic fields instead of AI extraction
            if "Abstract" in paper_df.columns:
                # Get abstract from paper details (fetched from PubMed)
                abstract_text = base_info.get("abstract", "No abstract available")
                paper_df["Abstract"] = abstract_text

            # If AI produced no rows from available full text, log it but do NOT create
            # an empty-gene placeholder row — it has zero informational value and would
            # appear in the researcher-facing CSV with a misleading MEDIUM confidence.
            if paper_df.empty:
                if not worker_debug:
                    worker_debug = {"status": "empty_result"}
            else:
                # Defensively ensure unique column names before concatenation
                paper_df = _ensure_unique_columns(paper_df)
                all_results_df = _ensure_unique_columns(all_results_df)
                all_results_df = pd.concat([all_results_df, paper_df], ignore_index=True)
                collected_rows.append(pmid)
                full_rows_pmids.add(pmid)
                # Update genes_extracted stat
                if "Gene/Group" in paper_df.columns:
                    unique_genes = paper_df["Gene/Group"].dropna().nunique()
                    pipeline_stats["genes_extracted"] = (
                        pipeline_stats.get("genes_extracted", 0) + unique_genes
                    )
                    if unique_genes > 0:
                        emit_log("info", f"Extracted {unique_genes} genes from PMID {pmid}")

            paper_debug_artifacts.append(
                {
                    "pmid": pmid,
                    "status": worker_debug.get("status", "ok"),
                    "reason": worker_debug.get("reason", ""),
                    "candidate_count": worker_debug.get("candidate_count"),
                    "candidates": worker_debug.get("candidates", []),
                    "detail_extraction_status": worker_debug.get("detail_extraction_status", ""),
                    "detail_extraction_error": worker_debug.get("detail_extraction_error", ""),
                    "detail_extraction_rows": worker_debug.get("detail_extraction_rows"),
                    "validation_drops": worker_debug.get("validation_drops", []),
                    "strict_gate_drops": worker_debug.get("strict_gate_drops", []),
                    "evidence_gate_drops": worker_debug.get("evidence_gate_drops", []),
                    "final_associations": worker_debug.get("final_associations", []),
                    "emitted_rows": int(len(paper_df)),
                }
            )

    except JobCancelledException:
        logging.warning("Job cancellation detected! Stopping new paper processing.")

        # Mark remaining PMIDs as Cancelled
        processed_set = set(collected_rows) | {r["PMID"] for r in minimal_rows}
        remaining = [p for p in pmids_to_process if p not in processed_set]

        for pmid in remaining:
            base_info = paper_details.get(pmid, {})
            minimal_rows.append(
                {
                    **build_minimal_row(
                        pmid,
                        base_info,
                        citation_records,
                        gene_group="CANCELLED",
                        variant_name="CANCELLED",
                    ),
                    "Abstract": base_info.get("abstract", "No abstract available - Job Cancelled"),
                }
            )
    finally:
        # Always clean up the worker pool (handles normal completion, cancellation, and errors)
        try:
            worker_pool.terminate()
            worker_pool.join(timeout=10)
        except Exception as e:
            logging.warning(f"Worker pool final cleanup failed: {e}")
        logging.info("AI worker pool terminated")

    if all_results_df.empty:
        if minimal_rows:
            logging.warning("AI produced no rows; falling back to minimal rows only.")
            all_results_df = pd.DataFrame(minimal_rows)
        else:
            logging.warning("No results were extracted by the pipeline.")
            report_progress("Completed", 100)
            return None

    # HYBRID PIPELINE: NCBI Gene enrichment
    if getattr(config, "ENABLE_NCBI_ENRICHMENT", True) and not all_results_df.empty:
        report_progress("Enriching gene metadata", 92)
        logging.info("STEP 5.5: Enriching genes with NCBI Gene metadata...")

        try:
            # Collect all unique genes that need enrichment
            genes_to_enrich = set()
            if "Gene/Group" in all_results_df.columns:
                genes_to_enrich = set(all_results_df["Gene/Group"].dropna().unique())
                genes_to_enrich = {g for g in genes_to_enrich if g}

            if genes_to_enrich:
                ncbi_tool = NCBIGeneTool()
                gene_metadata = ncbi_tool.enrich_gene_symbols(list(genes_to_enrich))

                # Add enrichment columns
                def get_full_name(gene):
                    if not gene or str(gene).upper() not in {
                        k.upper() for k in gene_metadata.keys()
                    }:
                        return ""
                    for k, v in gene_metadata.items():
                        if k.upper() == str(gene).upper():
                            return v.full_name
                    return ""

                def get_aliases(gene):
                    if not gene:
                        return ""
                    for k, v in gene_metadata.items():
                        if k.upper() == str(gene).upper():
                            return ", ".join(v.aliases[:3]) if v.aliases else ""
                    return ""

                def get_chromosome(gene):
                    if not gene:
                        return ""
                    for k, v in gene_metadata.items():
                        if k.upper() == str(gene).upper():
                            return v.chromosome or ""
                    return ""

                def get_ncbi_gene_id(gene):
                    if not gene:
                        return ""
                    for k, v in gene_metadata.items():
                        if k.upper() == str(gene).upper():
                            return v.gene_id or ""
                    return ""

                all_results_df["Gene Full Name"] = all_results_df["Gene/Group"].apply(get_full_name)
                all_results_df["Gene Aliases"] = all_results_df["Gene/Group"].apply(get_aliases)
                all_results_df["Chromosome"] = all_results_df["Gene/Group"].apply(get_chromosome)

                # Backfill NCBI Gene ID for rows that don't already have one (e.g. LLM-only genes
                # not found by PubTator). The enrichment API returns gene_id for every resolved symbol.
                if "NCBI Gene ID" not in all_results_df.columns:
                    all_results_df["NCBI Gene ID"] = ""
                missing_mask = (
                    all_results_df["NCBI Gene ID"].fillna("").astype(str).str.strip() == ""
                )
                all_results_df.loc[missing_mask, "NCBI Gene ID"] = all_results_df.loc[
                    missing_mask, "Gene/Group"
                ].apply(get_ncbi_gene_id)

                logging.info(
                    f"NCBI enrichment: Added metadata for {len(gene_metadata)}/{len(genes_to_enrich)} genes"
                )
        except Exception as e:
            logging.warning(f"NCBI enrichment failed, continuing without it: {e}")

    # Enforce desired output policy:
    # - Include up to 'desired' fully reviewed studies (AI-produced rows) first
    # - Append minimal rows for remaining scraped PMIDs at the end (do not count toward N)
    desired = top_n_cited
    if desired:
        # Order full-review PMIDs by processing order
        ordered_full = []
        for pmid in collected_rows:
            if pmid in full_rows_pmids and pmid not in ordered_full:
                ordered_full.append(pmid)
        selected_full_list = ordered_full[:desired]
        selected_full_set = set(selected_full_list)

        # Keep only rows for selected full-review PMIDs (up to N)
        full_df = all_results_df[all_results_df["PMID"].isin(selected_full_set)].reset_index(
            drop=True
        )

        # Build minimal rows for PMIDs that had no full rows (in original processing order)
        minimal_append = []
        selected_full_or_minimal_pmids = set(selected_full_set)
        for mr in minimal_rows:
            if mr["PMID"] in selected_full_or_minimal_pmids:
                continue
            minimal_append.append(mr)
            selected_full_or_minimal_pmids.add(mr["PMID"])

        if minimal_append:
            all_results_df = pd.concat([full_df, pd.DataFrame(minimal_append)], ignore_index=True)
        else:
            all_results_df = full_df

    report_progress("Finalizing results", 95)

    # De-duplicate: If multiple rows within the same PMID and gene have identical user fields, aggregate variant names
    try:
        if not all_results_df.empty:
            core_cols = [
                "PMID",
                "DOI",
                "Gene/Group",
                "Variant Name",
                "Study Title",
                "Authors",
                "Publication Year",
                "Journal Name",
                "Author Affiliations",
                "Citations",
                "Citation Source",
                "Citation Retrieved At",
                "iCite Citations",
                "Semantic Scholar Citations",
                "Figure Count",
                "Figure Analysis Enabled",
                "Candidate Source",
                "Normalization Applied",
                "Validation Outcome",
                "Dropped By Gate",
            ]
            # Identify user columns (including their citation pairs) that are not core or metadata
            metadata_suffixes = [
                "_citation_valid",
                "_citation_confidence",
                "_citation_details",
                "validation_confidence",
                "validation_source",
                "validation_suggestions",
                "context_flash_fits",
                "context_pro_fits",
                "context_original_tokens",
                "context_modifications",
                "context_truncation_applied",
            ]

            user_cols = [
                c
                for c in all_results_df.columns
                if c not in core_cols
                and not any(c.endswith(s) or c == s for s in metadata_suffixes)
            ]

            # Grouping keys exclude 'Variant Name' so we can aggregate it when other fields are identical
            group_keys = [
                "PMID",
                "Gene/Group",
                "Study Title",
                "Authors",
                "Publication Year",
                "Journal Name",
            ] + user_cols

            # Only dedup rows that actually have a gene (skip minimal rows lacking gene info)
            has_gene_mask = all_results_df["Gene/Group"].fillna("") != ""
            df_full = all_results_df[has_gene_mask].copy()
            df_minimal = all_results_df[~has_gene_mask].copy()

            if not df_full.empty:
                agg_map = {
                    c: "first" for c in df_full.columns if c not in group_keys + ["Variant Name"]
                }

                # Aggregate variant names: unique; sorted; semicolon-joined
                def _agg_variants(series):
                    vals = {str(v) for v in series if str(v).strip()}
                    return "; ".join(sorted(vals))

                agg_map["Variant Name"] = _agg_variants

                df_full = df_full.groupby(group_keys, dropna=False, as_index=False).agg(agg_map)

                # Recombine full and minimal, preserving original order later via sort key
                all_results_df = pd.concat([df_full, df_minimal], ignore_index=True)
    except Exception as e:
        logging.warning(f"Deduplication step skipped due to error: {e}")

    # Ensure unique columns before reordering to prevent reindex errors
    all_results_df = _ensure_unique_columns(all_results_df)

    # Reorder columns for better readability
    core_columns = [
        "Gene/Group",
        "Variant Name",
        "Gene Source",
        "NCBI Gene ID",
        "Gene Full Name",
        "Gene Aliases",
        "Gene Biotype",
        "Chromosome",
        "Candidate Source",
        "Normalization Applied",
        "Validation Outcome",
        "Dropped By Gate",
        "PMID",
        "DOI",
        "Study Title",
        "Authors",
        "Publication Year",
        "Journal Name",
        "Author Affiliations",
        "Citations",
        "Citation Source",
        "Citation Retrieved At",
        "iCite Citations",
        "Semantic Scholar Citations",
        "Figure Count",
        "Figure Analysis Enabled",
        "Metadata Completeness",
        "Metadata Warnings",
    ]

    # User-defined columns (exclude core and metadata columns)
    metadata_suffixes = [
        "_citation_valid",
        "_citation_confidence",
        "_citation_details",
        "validation_confidence",
        "validation_source",
        "validation_suggestions",
        "context_flash_fits",
        "context_pro_fits",
        "context_original_tokens",
        "context_modifications",
        "context_truncation_applied",
    ]

    user_columns_raw = [
        col
        for col in all_results_df.columns
        if col not in core_columns
        and not any(col.endswith(suffix) or col == suffix for suffix in metadata_suffixes)
    ]
    # Pair user columns with their citation columns if present
    user_columns_list = []
    used = set()
    for col in user_columns_raw:
        if col in used:
            continue
        citation_col = f"{col} Citation"
        user_columns_list.append(col)
        if citation_col in all_results_df.columns:
            user_columns_list.append(citation_col)
            used.add(citation_col)
        used.add(col)

    # Metadata columns at the end
    metadata_columns = [
        col
        for col in all_results_df.columns
        if any(col.endswith(suffix) or col == suffix for suffix in metadata_suffixes)
    ]

    # Build final column order
    final_column_order = []
    for col in core_columns:
        if col in all_results_df.columns:
            final_column_order.append(col)
    final_column_order.extend(user_columns_list)
    final_column_order.extend(metadata_columns)

    all_results_df = all_results_df[final_column_order]

    # Step 6: Save final results — primary CSV + metadata CSV + Excel + JSON
    output_path = _create_unique_filepath("final_enriched_results", "csv")
    primary_path, metadata_path, excel_path, json_path = _write_split_output(
        df=all_results_df,
        output_path=output_path,
        user_cols=user_columns_raw,
    )
    debug_path = write_drop_debug_artifact(status="completed", output_csv_path=primary_path)

    total_genes = pipeline_stats.get("genes_extracted", 0)
    total_analyzed = analyzed_attempts
    elapsed = time.time() - start_time
    emit_log(
        "info",
        f"Pipeline complete: {total_genes} genes from {total_analyzed} papers ({elapsed:.0f}s)",
    )
    report_progress("Completed", 100)
    logging.info(f"--- PIPELINE FINISHED IN {elapsed:.2f} SECONDS ---")
    return {
        "local_path": primary_path,
        "metadata_path": metadata_path,
        "excel_path": excel_path,
        "json_path": json_path,
        "debug_path": debug_path,
    }
