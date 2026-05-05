# modules/pipeline_orchestrator.py

import gzip
import json
import logging
import multiprocessing as mp
import os
import pickle
import time
import uuid

import pandas as pd
from tqdm import tqdm

from . import config, full_text_fetcher, pipeline_tracer, pubmed_data_collector
from .abstract_screener import has_genetic_content
from .association_policy import (
    ASSOCIATION_GROUP_ORDER,
    association_group_for_type,
    count_association_groups,
)
from .content_preparation import PreparedPaperContent
from .pubtator_tool import HybridExtractionResult, NCBIGeneTool, PubTatorTool
from .paper_analysis.pipeline import PaperAnalysisPipeline


class JobCancelledException(Exception):
    """Raised when a job is cancelled by the user."""

    pass


def _compute_row_confidence(row: dict, user_cols: list) -> tuple:
    """Compute a single confidence signal per output row.

    Returns (level, note) where level is HIGH/MEDIUM/LOW/REVIEW.

    - HIGH:   Gene corroborated by both PubTator NER + Gemini LLM, AND at least
              one citation verified in paper text. Strongest evidence tier.
    - MEDIUM: Gene passed all validation gates (confidence >= 0.85, grounding
              check, strict gate) but lacks full corroboration. Typical cases:
              LLM-only gene (PubTator missed it), corroborated gene without a
              validated citation (LLM stochastically omitted citations), or
              single-source gene with a valid citation.
    - LOW:    Abstract-only paper (no full text available) or validation
              confidence below 0.85 (borderline HGNC match).
    - REVIEW: Citation text not found in paper (mismatch), gene extracted
              only from figures with no prose source, OR row was produced by
              skeleton/auto-snippet fallback because Gemini failed (F11).
              Requires manual check.
    """
    # Guard: empty gene name is never a valid extraction
    if not str(row.get("Gene/Group", "") or "").strip():
        return "REVIEW", "No genes extracted"

    # F11: Extraction-mode check runs before any other tier logic — a skeleton
    # row can't be HIGH/MEDIUM/LOW even if the gene's HGNC validation is 1.0,
    # because the LLM didn't read the paper's context for this gene. The row
    # exists to preserve the gene identity, not as an actual extraction result.
    extraction_mode = str(row.get("extraction_mode", "") or "").strip()
    if extraction_mode == "skeleton":
        err = str(row.get("detail_extraction_error", "") or "").strip()
        err_short = (err[:80] + "…") if len(err) > 80 else err
        evidence_backfilled = bool(row.get("evidence_backfilled", False))
        if evidence_backfilled:
            note = "LLM failed; auto-snippet fallback" + (f" ({err_short})" if err_short else "")
        else:
            note = "LLM failed; no content" + (f" ({err_short})" if err_short else "")
        # F12: surface peer-gene context when the backfill snippet was a co-mention
        # rather than a gene-specific sentence.
        if row.get("evidence_specificity") == "co_mention":
            co_mentioned = str(row.get("co_mentioned_genes") or "").strip()
            if co_mentioned:
                note += f" | co-mention with {co_mentioned.replace(';', ', ')}"
        return "REVIEW", note

    val_conf = float(row.get("validation_confidence", 0) or 0)
    gene_source = str(row.get("Gene Source", "") or "")
    candidate_source = str(row.get("Candidate Source", "") or "")
    context_mods = str(row.get("context_modifications", "") or "")
    has_full_text = "no_oa_full_text" not in context_mods

    # Citation signals — check all user columns
    any_valid = False
    any_false = False
    all_empty = True
    for col in user_cols:
        citation_text = str(row.get(f"{col} Citation", "") or "")
        if citation_text:
            all_empty = False
        valid_flag = row.get(f"{col}_citation_valid")
        if valid_flag is True:
            any_valid = True
        elif valid_flag is False:
            any_false = True

    is_figure_only = (
        "llm_figure" in candidate_source
        and "llm_text" not in candidate_source
        and "pubtator" not in candidate_source
        and "deterministic_lexicon" not in candidate_source
    )

    # REVIEW: citation mismatch or figure-only source
    if (any_false and not any_valid) or is_figure_only:
        note = (
            "Figure-only gene — no prose citation available"
            if is_figure_only
            else "Citation text not found in paper"
        )
        return "REVIEW", note

    # LOW: no full text or borderline confidence
    if not has_full_text:
        return "LOW", "Abstract only"
    if val_conf < 0.85:
        return "LOW", "Low confidence"

    # HIGH: corroborated by multiple sources AND verified citation
    if gene_source == "both" and any_valid:
        return "HIGH", ""

    # MEDIUM: default (passed gates, citation stochastic or absent)
    if all_empty:
        note = ""
    elif not any_valid:
        note = "No citation"
    else:
        note = ""
    # F12: defensively surface peer-gene context when a backfilled row lands
    # outside the skeleton branch (e.g. LLM returned content but evidence
    # gate pulled a co-mention snippet). Guarded on both flags so future
    # writers of evidence_specificity can't trigger this branch without the
    # causal chain — the only current writer is _backfill_sparse_row_evidence.
    if row.get("evidence_backfilled") and row.get("evidence_specificity") == "co_mention":
        co_mentioned = str(row.get("co_mentioned_genes") or "").strip()
        if co_mentioned:
            suffix = f"co-mention with {co_mentioned.replace(';', ', ')}"
            note = f"{note} | {suffix}" if note else suffix
    return "MEDIUM", note


logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - [%(module)s] - %(message)s"
)


def _run_pipeline_worker(
    text,
    cols,
    pubtator_genes=None,
    figure_inputs=None,
    abstract_text=None,
    table_inputs=None,
    pmid=None,
    prepared_content=None,
    pipeline_factory=PaperAnalysisPipeline,
):
    """Top-level worker function for multiprocessing pool (must be picklable).

    Returns the result dict directly; mp.Pool.apply_async transmits it back to
    the orchestrator via IPC, eliminating the need for an explicit Queue.

    Args:
        text: Full paper text
        cols: User-defined column descriptions
        pubtator_genes: Optional list of gene symbols from PubTator (high-precision NER)
        figure_inputs: Optional list of figure metadata dicts for Gemini multimodal analysis
        abstract_text: Abstract text for independent abstract-level gene discovery
        table_inputs: Optional list of structured table dicts for table-cell citation validation
        pmid: Optional PMID — used only by the pipeline tracer to decide whether
              to record detailed per-stage events for this paper.
        prepared_content: Optional PreparedPaperContent with normalized text/table indexes.
    """
    # Install the function-level tracer INSIDE the worker process too.
    # Without this, the live-stream function trace goes silent for the whole
    # Gemini phase (abstract pass, full-text passes, figure analysis, gates,
    # detail extraction, citation validation) because workers are separate
    # processes that don't inherit the parent's sys.setprofile hook.
    # install_function_tracer is idempotent and no-op when env vars are unset,
    # so this is free in untraced runs.
    trace_this_paper = pipeline_tracer.matches(pmid)
    try:
        if trace_this_paper:
            pipeline_tracer.install_function_tracer()
    except Exception:
        pass

    try:
        with pipeline_tracer.paper(pmid if trace_this_paper else None):
            inst = pipeline_factory(
                text,
                abstract_text=abstract_text or "",
                pubtator_genes=pubtator_genes,
                figure_inputs=figure_inputs,
                table_inputs=table_inputs,
                pmid=pmid,
                prepared_content=prepared_content,
            )
            df = inst.run_pipeline(cols)
            df = _ensure_unique_columns(df)
            return {
                "records": df.to_dict(orient="records"),
                "debug": inst._collect_debug_artifact(),
                "gemini_api_calls": inst._paper_api_calls,
            }
    except Exception as e:
        import traceback

        logging.error(f"Pipeline worker error: {e}\n{traceback.format_exc()}")
        return {"error": str(e)}
    finally:
        # Flush any tracer events this worker recorded before returning, even
            # when per-paper extraction raised after emitting partial trace events.
        try:
            pipeline_tracer.flush_worker_partial()
        except Exception:
            pass
        try:
            pipeline_tracer.uninstall_function_tracer()
        except Exception:
            pass


def _create_unique_filepath(filename_base, extension):
    unique_id = uuid.uuid4().hex[:8]
    filename = f"{filename_base}_{unique_id}.{extension}"
    return os.path.join(config.OUTPUT_DIR, filename)


def _write_excel_output(
    df_clean: "pd.DataFrame",
    df_meta: "pd.DataFrame",
    excel_path: "os.PathLike",
) -> None:
    """Write a two-sheet Excel workbook: Results (clean) + Metadata (full).

    Sheet 1 'Results': clean researcher-facing columns with Confidence color coding.
    Sheet 2 'Metadata': full diagnostic/validation columns for auditing.
    Silently skips if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logging.warning(
            "openpyxl not installed — Excel output skipped. "
            "Install with: pip install openpyxl"
        )
        return

    _CONFIDENCE_FILLS = {
        "HIGH": "D4EDDA",
        "MEDIUM": "FFF9C4",
        "LOW": "FFE0B2",
        "REVIEW": "FCE4EC",
    }

    def _fill_sheet(ws, df: "pd.DataFrame", sheet_title: str, apply_conf_color: bool) -> None:
        ws.title = sheet_title
        headers = list(df.columns)
        conf_col_idx = (headers.index("Confidence") + 1) if (apply_conf_color and "Confidence" in headers) else None
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=False)
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                if conf_col_idx and col_idx == conf_col_idx:
                    hex_color = _CONFIDENCE_FILLS.get(str(value))
                    if hex_color:
                        cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)
        ws.freeze_panes = "A2"
        # Auto-size columns (sample first 100 rows for speed)
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            sample_rows = range(2, min(ws.max_row + 1, 102))
            max_len = max(
                len(str(header)),
                max((len(str(ws.cell(row=r, column=col_idx).value or "")) for r in sample_rows), default=0),
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb = Workbook()
    _fill_sheet(wb.active, df_clean, "Results", apply_conf_color=True)
    _fill_sheet(wb.create_sheet("Metadata"), df_meta, "Metadata", apply_conf_color=False)
    if "Association Group" in df_clean.columns:
        for group, group_df in df_clean.groupby("Association Group", dropna=False, sort=False):
            group_name = str(group or "Unclassified")
            safe_title = "".join(ch for ch in group_name if ch not in r'[]:*?/\\')[:31]
            if safe_title and safe_title not in wb.sheetnames:
                _fill_sheet(
                    wb.create_sheet(safe_title),
                    group_df,
                    safe_title,
                    apply_conf_color=True,
                )
    wb.save(excel_path)


def _write_json_output(df_clean: "pd.DataFrame", json_path: "os.PathLike") -> None:
    """Write the clean results as a JSON array of records."""
    df_clean = _ensure_unique_columns(df_clean.copy())
    df_clean.to_json(json_path, orient="records", indent=2, force_ascii=False)


def _unique_preserve_order(items: list) -> list:
    """Return items with duplicates removed while preserving first appearance."""
    seen = set()
    unique = []
    for item in items:
        if item in seen:
            continue
        unique.append(item)
        seen.add(item)
    return unique


def _write_split_output(
    df: "pd.DataFrame",
    output_path: "os.PathLike",
    user_cols: list,
) -> tuple:
    """Write primary CSV, metadata CSV, Excel workbook, and JSON file.

    Primary CSV: Gene, Variant, [user cols], Confidence, Confidence Note,
                 PMID, Title, Year, Journal, Authors, Citations, DOI
    Metadata CSV: PMID, Gene/Group, Variant Name + all diagnostic/validation columns
    Excel: two-sheet workbook (Results + Metadata), Confidence cells color-coded
    JSON: primary CSV data as an array of records

    Both CSV files share PMID + Gene/Group as join keys and have identical row order.

    Args:
        df: Full results DataFrame with all columns.
        output_path: Destination path for the primary CSV (.csv extension).
        user_cols: List of user-defined column names (without Citation pairs).

    Returns:
        (primary_path, metadata_path, excel_path, json_path) as strings.
    """
    from pathlib import Path

    output_path = Path(output_path)
    df = _ensure_unique_columns(df.copy())

    # Compute confidence flags row-by-row
    confidence_rows = [
        _compute_row_confidence(row, user_cols)
        for row in df.to_dict("records")
    ]
    df = df.copy()
    df["Confidence"] = [r[0] for r in confidence_rows]
    df["Confidence Note"] = [r[1] for r in confidence_rows]
    if "Association Type" in df.columns and "Association Group" not in df.columns:
        df["Association Group"] = df["Association Type"].apply(association_group_for_type)
    if "Association Group" in df.columns:
        df["_association_group_order"] = df["Association Group"].map(
            lambda group: ASSOCIATION_GROUP_ORDER.get(str(group or ""), 99)
        )
        sort_cols = [
            col
            for col in ["_association_group_order", "Confidence", "PMID", "Gene/Group"]
            if col in df.columns
        ]
        df = df.sort_values(by=sort_cols, kind="stable").drop(
            columns=["_association_group_order"]
        )

    # --- Primary CSV (clean researcher-facing) ---
    rename_map = {
        "Gene/Group": "Gene",
        "Variant Name": "Variant",
        "Study Title": "Title",
        "Publication Year": "Year",
        "Journal Name": "Journal",
    }
    df_clean = df.rename(columns=rename_map)
    df_clean = _ensure_unique_columns(df_clean)

    primary_cols = (
        ["Gene", "Variant"]
        + [c for c in user_cols if c in df_clean.columns]
        + [
            "Confidence", "Confidence Note", "Association Group", "Association Type",
            # F11: extraction_mode (llm / skeleton) and evidence_backfilled
            # visible in the primary CSV so researchers can see fallback
            # rows at a glance instead of having to consult the metadata CSV.
            "extraction_mode", "evidence_backfilled", "evidence_specificity",
            "context_modifications",  # what sections were truncated per gene row
        ]
        + ["PMID", "Title", "Year", "Journal", "Authors", "Citations", "DOI"]
    )
    primary_cols_present = _unique_preserve_order([c for c in primary_cols if c in df_clean.columns])
    df_primary = df_clean[primary_cols_present]
    df_primary = _ensure_unique_columns(df_primary)
    df_primary.to_csv(output_path, index=False)

    # --- Metadata CSV (full transparency, same row order as primary) ---
    meta_path = output_path.with_name(output_path.stem + "_metadata.csv")
    join_keys = ["PMID", "Gene/Group", "Variant Name"]
    remaining = [c for c in df.columns if c not in join_keys]
    meta_cols_present = _unique_preserve_order([c for c in join_keys + remaining if c in df.columns])
    df_meta = df[meta_cols_present]
    df_meta = _ensure_unique_columns(df_meta)
    df_meta.to_csv(meta_path, index=False)

    # --- Excel workbook (Results + Metadata sheets) ---
    excel_path = output_path.with_suffix(".xlsx")
    _write_excel_output(df_primary, df_meta, excel_path)

    # --- JSON (primary data as records) ---
    json_path = output_path.with_suffix(".json")
    _write_json_output(df_primary, json_path)

    return str(output_path), str(meta_path), str(excel_path), str(json_path)


def _candidate_audit_rows_by_pmid(df: "pd.DataFrame") -> dict:
    """Summarize emitted final rows for candidate-audit final counts."""
    if df is None or df.empty:
        return {}

    work = _ensure_unique_columns(df.copy())
    if "Association Type" in work.columns and "Association Group" not in work.columns:
        work["Association Group"] = work["Association Type"].apply(association_group_for_type)

    pmid_col = "PMID" if "PMID" in work.columns else None
    gene_col = "Gene/Group" if "Gene/Group" in work.columns else ("Gene" if "Gene" in work.columns else None)
    variant_col = (
        "Variant Name"
        if "Variant Name" in work.columns
        else ("Variant" if "Variant" in work.columns else None)
    )
    if not pmid_col or not gene_col:
        return {}

    rows_by_pmid = {}
    for pmid, group_df in work.groupby(pmid_col, dropna=False, sort=False):
        associations = []
        for _, row in group_df.iterrows():
            gene = str(row.get(gene_col) or "")
            if not gene.strip():
                continue
            association_type = str(row.get("Association Type") or "")
            association_group = str(row.get("Association Group") or "")
            if not association_group:
                association_group = association_group_for_type(association_type)
            associations.append(
                {
                    "gene": gene,
                    "variant": str(row.get(variant_col) or "") if variant_col else "",
                    "association_type": association_type,
                    "association_group": association_group,
                }
            )
        rows_by_pmid[str(pmid)] = {
            "final_associations": associations,
            "emitted_rows": int(len(group_df)),
            "final_association_group_counts": count_association_groups(associations),
        }
    return rows_by_pmid


def _candidate_audit_summary(papers: list[dict]) -> dict:
    """Build run-level audit counts from emitted final rows."""
    final_group_counts = count_association_groups(
        association
        for paper in papers
        for association in (paper.get("final_associations") or [])
    )
    candidate_group_counts = count_association_groups(
        candidate for paper in papers for candidate in (paper.get("candidates") or [])
    )
    return {
        "papers": len(papers),
        "total_candidates": sum(int(p.get("candidate_count") or 0) for p in papers),
        "total_emitted_rows": sum(int(p.get("emitted_rows") or 0) for p in papers),
        "association_group_counts": final_group_counts,
        "final_association_group_counts": final_group_counts,
        "candidate_association_group_counts": candidate_group_counts,
    }


def _sanitize_user_columns(user_columns: dict) -> dict:
    """Rename user-provided columns that collide with reserved/core names.

    Ensures the keys sent to the extractor are unique and do not overlap with
    core columns like 'Gene/Group', 'PMID', etc. If a collision is detected,
    the column is renamed to 'User: <name>'.
    """
    if not user_columns:
        return {}

    reserved_core = {
        "Gene/Group",
        "Variant Name",
        "PMID",
        "Study Title",
        "Authors",
        "Publication Year",
        "Journal Name",
        "Author Affiliations",
        "Citations",
        "Abstract",
        # Primary output names after researcher-facing renames.
        "Gene",
        "Variant",
        "Title",
        "Year",
        "Journal",
        "Confidence",
        "Confidence Note",
        "Association Group",
        "Association Type",
        # Internal identifiers occasionally surfaced
        "gene_name",
        "variant_name",
    }

    out: dict = {}
    used: set = set()
    for raw_key, desc in user_columns.items():
        key = (raw_key or "").strip()
        if not key:
            continue
        # Avoid direct collisions with reserved core or already-used keys
        candidate = key
        if candidate in reserved_core or candidate in used:
            candidate = f"User: {candidate}"
        # If still collides, append numeric suffix until unique
        suffix = 2
        while candidate in reserved_core or candidate in used:
            candidate = f"{key} ({suffix})"
            suffix += 1
        out[candidate] = desc
        used.add(candidate)
    return out


def _ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure DataFrame has uniquely named columns by suffixing duplicates.

    Pandas reindexing requires unique column names; this function disambiguates
    duplicates by appending " (2)", "(3)", ... in appearance order.
    """
    if not isinstance(df, pd.DataFrame):
        return df
    cols = list(df.columns)
    seen_counts: dict[str, int] = {}
    new_cols: list[str] = []
    for c in cols:
        count = seen_counts.get(c, 0)
        if count == 0:
            new_cols.append(c)
        else:
            new_cols.append(f"{c} ({count + 1})")
        seen_counts[c] = count + 1
    df.columns = new_cols
    return df


def _get_citation_record(pmid, citation_records):
    rec = citation_records.get(pmid, {}) if citation_records else {}
    return {
        "count": int(rec.get("count", 0) or 0),
        "source": rec.get("source", "none") or "none",
        "retrieved_at": rec.get("retrieved_at", "") or "",
        "icite_count": rec.get("icite_count"),
        "semantic_scholar_count": rec.get("semantic_scholar_count"),
    }


def _is_gemini_quota_error(text: object) -> bool:
    err = str(text or "")
    return "429" in err or "RESOURCE_EXHAUSTED" in err or "quota" in err.lower()


def _prepare_paper_inputs(pmid, content_dict, paper_details, pubtator_results):
    content = content_dict.get(pmid, {})
    paper_text = content.get("content", "")
    figure_inputs = (
        content.get("figures", [])
        if getattr(config, "ENABLE_FIGURE_ANALYSIS", True)
        else []
    )
    table_inputs = (
        content.get("tables", [])
        if getattr(config, "ENABLE_TABLE_CITATIONS", True)
        else []
    )
    base_info = paper_details.get(pmid, {})
    abstract = base_info.get("abstract", "")
    title = base_info.get("title", "")

    pt_gene_symbols = []
    if pmid in pubtator_results:
        pt_gene_symbols = [g.symbol for g in pubtator_results[pmid].pubtator_genes]
        if pt_gene_symbols:
            logging.debug(
                f"PMID {pmid}: Passing {len(pt_gene_symbols)} PubTator genes to LLM"
            )

    return {
        "pmid": pmid,
        "paper_text": paper_text,
        "figure_inputs": figure_inputs,
        "table_inputs": table_inputs,
        "prepared_content": PreparedPaperContent.from_raw(
            paper_text=paper_text,
            abstract_text=abstract,
            table_inputs=table_inputs,
        ),
        "base_info": base_info,
        "abstract": abstract,
        "title": title,
        "pt_gene_symbols": pt_gene_symbols,
    }


def _aggregate_strict_gate_drops(
    pipeline_stats,
    worker_debug,
    pmid,
):
    """
    Aggregate per-paper strict-gate drops into the run-level ``pipeline_stats``.

    F10b helper — pure function, no side-effects beyond mutating
    ``pipeline_stats``. Each drop dict from ``worker_debug["strict_gate_drops"]``
    is copied, tagged with the given ``pmid``, and appended to the run-level
    ``strict_gate_drops`` list; the ``strict_gate_drops_count`` counter is
    refreshed from the list length.

    Extracted from ``_finalize_paper_result`` to enable direct unit testing
    without building the full worker payload and pandas stack.
    """
    paper_drops = worker_debug.get("strict_gate_drops", []) or []
    for drop in paper_drops:
        drop_entry = dict(drop)
        drop_entry["pmid"] = pmid
        pipeline_stats.setdefault("strict_gate_drops", []).append(drop_entry)
    pipeline_stats["strict_gate_drops_count"] = len(
        pipeline_stats.get("strict_gate_drops", [])
    )
    return paper_drops


def _gene_key(gene):
    return str(gene or "").strip().upper()


def _apply_pubtator_row_enrichment(paper_df, pmid, pubtator_results):
    """Attach PubTator source and NCBI IDs with per-paper lookup caches."""
    if paper_df.empty or pmid not in pubtator_results or "Gene/Group" not in paper_df.columns:
        return paper_df

    hybrid_result = pubtator_results[pmid]
    pt_by_symbol = {
        _gene_key(g.symbol): g
        for g in hybrid_result.pubtator_genes
        if _gene_key(g.symbol)
    }
    llm_symbols = {
        _gene_key(g)
        for g in paper_df["Gene/Group"].dropna().unique().tolist()
        if _gene_key(g)
    }
    hybrid_result.llm_genes = sorted(llm_symbols)

    source_by_symbol = {}
    ncbi_id_by_symbol = {}
    for symbol in llm_symbols:
        source_by_symbol[symbol] = "both" if symbol in pt_by_symbol else "llm"
        ncbi_id_by_symbol[symbol] = getattr(pt_by_symbol.get(symbol), "ncbi_gene_id", "") or ""

    paper_df["Gene Source"] = paper_df["Gene/Group"].map(
        lambda gene: source_by_symbol.get(_gene_key(gene), "")
    )
    paper_df["NCBI Gene ID"] = paper_df["Gene/Group"].map(
        lambda gene: ncbi_id_by_symbol.get(_gene_key(gene), "")
    )
    return paper_df


def _apply_ncbi_metadata_columns(all_results_df, gene_metadata):
    """Attach NCBI enrichment columns using a single uppercase-keyed lookup."""
    if all_results_df.empty or "Gene/Group" not in all_results_df.columns:
        return all_results_df

    metadata_by_symbol = {
        _gene_key(symbol): value
        for symbol, value in (gene_metadata or {}).items()
        if _gene_key(symbol)
    }

    all_results_df["Gene Full Name"] = all_results_df["Gene/Group"].map(
        lambda gene: getattr(metadata_by_symbol.get(_gene_key(gene)), "full_name", "") or ""
    )
    all_results_df["Gene Aliases"] = all_results_df["Gene/Group"].map(
        lambda gene: ", ".join((getattr(metadata_by_symbol.get(_gene_key(gene)), "aliases", None) or [])[:3])
    )
    all_results_df["Chromosome"] = all_results_df["Gene/Group"].map(
        lambda gene: getattr(metadata_by_symbol.get(_gene_key(gene)), "chromosome", "") or ""
    )

    if "NCBI Gene ID" not in all_results_df.columns:
        all_results_df["NCBI Gene ID"] = ""
    missing_mask = (
        all_results_df["NCBI Gene ID"].fillna("").astype(str).str.strip() == ""
    )
    all_results_df.loc[missing_mask, "NCBI Gene ID"] = all_results_df.loc[
        missing_mask, "Gene/Group"
    ].map(lambda gene: getattr(metadata_by_symbol.get(_gene_key(gene)), "gene_id", "") or "")
    return all_results_df


def _finalize_paper_result(
    payload,
    pmid,
    base_info,
    citation_records,
    figure_inputs,
    pubtator_results,
    pipeline_stats,
    emit_log,
):
    """Normalize worker output, attach metadata, and build the debug artifact."""
    paper_df = pd.DataFrame()
    worker_debug = {}

    if isinstance(payload, dict) and payload.get("error"):
        logging.error(f"AI analysis error for PMID {pmid}: {payload['error']}")
        worker_debug = {
            "status": "worker_error",
            "reason": str(payload.get("error")),
        }
    elif isinstance(payload, dict) and payload.get("records") is not None:
        paper_df = pd.DataFrame(payload["records"])
        if isinstance(payload.get("debug"), dict):
            worker_debug = payload["debug"]
        else:
            worker_debug = {"status": "ok"}
        pipeline_stats["gemini_api_calls"] += int(payload.get("gemini_api_calls", 0))
        ctx_warn = (worker_debug or {}).get("context_warning")
        if ctx_warn:
            emit_log("warn", f"PMID {pmid}: {ctx_warn}")

    paper_df = paper_df.rename(
        columns={"gene_name": "Gene/Group", "variant_name": "Variant Name"}
    )

    paper_df["PMID"] = pmid
    paper_df["DOI"] = base_info.get("doi", "")
    paper_df["Study Title"] = base_info.get("title", "N/A")
    paper_df["Authors"] = ", ".join(base_info.get("authors", []))
    paper_df["Publication Year"] = base_info.get("year", "N/A")
    paper_df["Journal Name"] = base_info.get("journal", "N/A")
    paper_df["Author Affiliations"] = "; ".join(base_info.get("affiliations", []))
    citation = _get_citation_record(pmid, citation_records)
    paper_df["Citations"] = citation["count"]
    paper_df["Citation Source"] = citation["source"]
    paper_df["Citation Retrieved At"] = citation["retrieved_at"]
    paper_df["iCite Citations"] = citation["icite_count"]
    paper_df["Semantic Scholar Citations"] = citation["semantic_scholar_count"]
    paper_df["Metadata Completeness"] = base_info.get("_metadata_completeness", 0)
    paper_df["Metadata Warnings"] = "; ".join(base_info.get("_metadata_warnings", []))
    paper_df["Figure Count"] = len(figure_inputs)
    paper_df["Figure Analysis Enabled"] = bool(
        getattr(config, "ENABLE_FIGURE_ANALYSIS", True)
    )

    paper_df = _apply_pubtator_row_enrichment(paper_df, pmid, pubtator_results)

    if "Abstract" in paper_df.columns:
        paper_df["Abstract"] = base_info.get("abstract", "No abstract available")

    if paper_df.empty and not worker_debug:
        worker_debug = {"status": "empty_result"}

    detail_error = worker_debug.get("detail_extraction_error", "")
    detail_status = worker_debug.get("detail_extraction_status", "")
    quota_limited = (
        bool(worker_debug.get("quota_limited"))
        or _is_gemini_quota_error(detail_error)
        or detail_status == "quota_limited_fallback"
    )
    if quota_limited:
        row_count = int(len(paper_df))
        pipeline_stats["quota_limited_papers"] = int(
            pipeline_stats.get("quota_limited_papers", 0)
        ) + 1
        pipeline_stats["quota_limited_rows"] = int(
            pipeline_stats.get("quota_limited_rows", 0)
        ) + row_count
        emit_log(
            "warn",
            f"PMID {pmid}: Gemini quota/rate limit reached; output rows require review",
            "Rows were saved as fallback skeletons with auto-snippets, not full AI extraction.",
        )

    # F10b: Aggregate strict-gate drops into run-level pipeline_stats so the
    # Results UI can surface a banner without the operator opening the
    # drop_debug_*.json forensic file. Each entry is tagged with its PMID.
    paper_drops = _aggregate_strict_gate_drops(pipeline_stats, worker_debug, pmid)

    debug_artifact = {
        "pmid": pmid,
        "status": worker_debug.get("status", "ok"),
        "reason": worker_debug.get("reason", ""),
        "candidate_count": worker_debug.get("candidate_count"),
        "candidates": worker_debug.get("candidates", []),
        "detail_extraction_status": worker_debug.get("detail_extraction_status", ""),
        "detail_extraction_error": worker_debug.get("detail_extraction_error", ""),
        "quota_limited": quota_limited,
        "detail_extraction_rows": worker_debug.get("detail_extraction_rows"),
        "validation_drops": worker_debug.get("validation_drops", []),
        "strict_gate_drops": paper_drops,
        "evidence_gate_drops": worker_debug.get("evidence_gate_drops", []),
        "final_associations": worker_debug.get("final_associations", []),
        "emitted_rows": int(len(paper_df)),
    }
    return paper_df, debug_artifact


def _accumulate_result(
    all_results_df,
    paper_df,
    pmid,
    collected_rows,
    full_rows_pmids,
    pipeline_stats,
    emit_log,
):
    if paper_df.empty:
        return all_results_df

    paper_df = _ensure_unique_columns(paper_df)
    all_results_df = _ensure_unique_columns(all_results_df)
    all_results_df = pd.concat([all_results_df, paper_df], ignore_index=True)
    collected_rows.append(pmid)
    full_rows_pmids.add(pmid)

    if "Gene/Group" in paper_df.columns:
        unique_genes = paper_df["Gene/Group"].dropna().nunique()
        pipeline_stats["genes_extracted"] = (
            pipeline_stats.get("genes_extracted", 0) + unique_genes
        )
        if unique_genes > 0:
            emit_log("info", f"Extracted {unique_genes} genes from PMID {pmid}")

    return all_results_df


def run_complete_pipeline(
    query,
    specific_pmids,
    specific_authors,
    user_columns,
    top_n_cited,
    progress_callback=None,
    log_callback=None,
):
    """
    Runs the complete pipeline with the provided parameters.

    Args:
        log_callback: Optional callback(level, msg, detail) for structured log events.
                      Levels: "info" (user-facing), "debug" (technical), "warn", "error"
    """
    logging.info("--- STARTING COMPLETE PIPELINE RUN ---")
    start_time = time.time()

    def emit_log(level, msg, detail=None):
        """Emit a structured log event to the frontend."""
        if log_callback:
            log_callback(level, msg, detail)
        # Also route to Python logging for stderr capture
        log_level = {
            "info": logging.INFO,
            "debug": logging.DEBUG,
            "warn": logging.WARNING,
            "error": logging.ERROR,
        }.get(level, logging.INFO)
        logging.log(log_level, msg)

    def check_cancellation():
        """Check if job was cancelled via local .cancel file."""
        cancel_path = os.path.join(config.OUTPUT_DIR, ".cancel")
        if os.path.exists(cancel_path):
            emit_log("warn", "Pipeline cancelled by user")
            raise JobCancelledException("Job was cancelled")

    # Track pipeline statistics for frontend display
    pipeline_stats = {
        "papers_found": 0,
        "papers_screened": 0,
        "papers_screened_passed": 0,
        "papers_fetch_failed": 0,
        "papers_analyzed": 0,
        "genes_extracted": 0,
        "gemini_api_calls": 0,
        "tables_extracted": 0,
        "pubtator_pmids_skipped": 0,
        "papers_excluded_not_oa": 0,  # F2: count of paywalled PMIDs dropped at the backup gate
        "strict_gate_drops": [],  # F10b: flat list of per-gene strict-gate drops across all papers
        "strict_gate_drops_count": 0,  # F10b: operator-facing count surfaced to Results UI
        "quota_limited_papers": 0,
        "quota_limited_rows": 0,
    }
    paper_debug_artifacts = []
    forensic_screening = []
    fetch_report = []

    def build_minimal_row(pmid, base_info, citation_records, gene_group="", variant_name=""):
        citation = _get_citation_record(pmid, citation_records)
        return {
            "Gene/Group": gene_group,
            "Variant Name": variant_name,
            "Candidate Source": "",
            "Normalization Applied": "",
            "Validation Outcome": "",
            "Association Type": "",
            "Association Group": "Review Needed",
            "Dropped By Gate": "",
            "PMID": pmid,
            "DOI": base_info.get("doi", ""),
            "Study Title": base_info.get("title", "N/A"),
            "Authors": ", ".join(base_info.get("authors", [])),
            "Publication Year": base_info.get("year", "N/A"),
            "Journal Name": base_info.get("journal", "N/A"),
            "Author Affiliations": "; ".join(base_info.get("affiliations", [])),
            "Citations": citation["count"],
            "Citation Source": citation["source"],
            "Citation Retrieved At": citation["retrieved_at"],
            "iCite Citations": citation["icite_count"],
            "Semantic Scholar Citations": citation["semantic_scholar_count"],
            "Figure Count": 0,
            "Figure Analysis Enabled": bool(getattr(config, "ENABLE_FIGURE_ANALYSIS", True)),
            "Abstract": base_info.get("abstract", "No abstract available"),
            "Metadata Completeness": base_info.get("_metadata_completeness", 0),
            "Metadata Warnings": "; ".join(base_info.get("_metadata_warnings", [])),
        }

    def report_progress(stage, pct, extra_stats=None):
        check_cancellation()
        # Merge extra stats into pipeline_stats
        if extra_stats:
            pipeline_stats.update(extra_stats)
        if progress_callback:
            progress_callback(stage, pct, pipeline_stats.copy())
        logging.info(f"Progress: {stage} ({pct}%) - Stats: {pipeline_stats}")

    def write_drop_debug_artifact(status: str, output_csv_path: str = ""):
        """
        Persist per-run candidate/drop diagnostics to JSON for trust debugging.

        Includes forensic data when available: screening decisions, fetch outcomes,
        and table extraction statistics.
        """
        payload = {
            "status": status,
            "generated_at_epoch": time.time(),
            "query": query,
            "specific_pmids": list(specific_pmids or []),
            "specific_authors": list(specific_authors or []),
            "top_n_cited": top_n_cited,
            "output_csv_path": output_csv_path or "",
            "paper_debug": paper_debug_artifacts,
            "pipeline_stats": pipeline_stats.copy(),
            "screening_decisions": forensic_screening,
            "fetch_outcomes": fetch_report,
        }
        debug_path = _create_unique_filepath("drop_debug", "json")
        try:
            with open(debug_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
            emit_log("info", "Saved drop-debug artifact", debug_path)
            return debug_path
        except Exception as e:
            logging.warning(f"Failed to write drop-debug artifact: {e}")
            return ""

    def write_candidate_audit_artifact(output_csv_path: str = ""):
        """
        Persist candidate lifecycle as a stable first-class artifact.

        This is a cleaner consumer-facing subset of drop_debug: one paper record
        per PMID with candidate provenance, gate drops, and final associations.
        """
        emitted_by_pmid = _candidate_audit_rows_by_pmid(all_results_df)
        papers = []
        for paper in paper_debug_artifacts:
            emitted_summary = emitted_by_pmid.get(str(paper.get("pmid", "")), {})
            candidates = paper.get("candidates", []) or []
            final_associations = (
                emitted_summary.get("final_associations")
                if emitted_summary
                else paper.get("final_associations", []) or []
            )
            papers.append(
                {
                    "pmid": paper.get("pmid", ""),
                    "status": paper.get("status", ""),
                    "reason": paper.get("reason", ""),
                    "candidate_count": paper.get("candidate_count"),
                    "candidates": candidates,
                    "validation_drops": paper.get("validation_drops", []),
                    "strict_gate_drops": paper.get("strict_gate_drops", []),
                    "evidence_gate_drops": paper.get("evidence_gate_drops", []),
                    "final_associations": final_associations,
                    "emitted_rows": emitted_summary.get(
                        "emitted_rows", paper.get("emitted_rows", 0)
                    ),
                    "association_group_counts": count_association_groups(candidates),
                    "final_association_group_counts": emitted_summary.get(
                        "final_association_group_counts",
                        count_association_groups(final_associations),
                    ),
                }
            )
        payload = {
            "schema_version": "candidate_audit_v1",
            "status": "completed",
            "generated_at_epoch": time.time(),
            "output_csv_path": output_csv_path or "",
            "summary": _candidate_audit_summary(papers),
            "papers": papers,
        }
        audit_path = _create_unique_filepath("candidate_audit", "json")
        try:
            with open(audit_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
            emit_log("info", "Saved candidate audit artifact", audit_path)
            return audit_path
        except Exception as e:
            logging.warning(f"Failed to write candidate audit artifact: {e}")
            return ""

    report_progress("Initializing pipeline", 0)
    emit_log("info", "Initializing pipeline")

    # If tracing is enabled, point the tracer at the output dir so workers write partials there
    if pipeline_tracer.is_enabled():
        pipeline_tracer.set_output_dir(os.path.join(config.OUTPUT_DIR, ".trace_partials"))
        emit_log("info", f"Pipeline tracing enabled for PMID {pipeline_tracer.target_pmid()}")
        # Function-level tracing is installed here and inside each worker. Stage
        # contexts add stage_id to call/return events so the viewer can group
        # noisy function streams by semantic pipeline stage.
        if pipeline_tracer.function_tracer_enabled():
            pipeline_tracer.install_function_tracer()
            emit_log("info", "Function-level tracing active (orchestrator process)")

    # Step 1 & 2: Get PMIDs and Details
    report_progress("Searching PubMed", 10)
    if query:
        emit_log("info", f"Searching PubMed for: {query}")
    if specific_pmids:
        emit_log("info", f"Including {len(specific_pmids)} specific PMIDs")
    if specific_authors:
        emit_log("info", f"Searching papers by {len(specific_authors)} author(s)")
    logging.info("STEP 1 & 2: Gathering and fetching details for all PMIDs...")

    initial_pmids = set()
    mandatory_pmids = set(specific_pmids)  # Start with explicit PMIDs
    author_pmids = set()

    if specific_authors:
        for author in specific_authors:
            check_cancellation()
            author_results = pubmed_data_collector.search_pubmed_by_author(author, max_results=200)
            initial_pmids.update(author_results)
            author_pmids.update(author_results)

    mandatory_pmids |= author_pmids  # Union all author-derived PMIDs as mandatory

    # ── F2: OA backup gate for specific-PMIDs / author-search paths ──
    # ResearchShop is OA-only. The UI's SmartInput enforces this upstream
    # (green "Full text" / amber "No OA full text" badge). This gate is
    # defence-in-depth for CLI users and scripted invocations that bypass the UI.
    # Query-mode papers already pass through PubMed's `loattrfull text[sb]`
    # filter — they're unaffected here.
    papers_excluded_not_oa = 0
    if mandatory_pmids:
        check_cancellation()
        # Single-PMID PMC lookup via the fetcher's existing helper (same call it
        # would make later anyway). Typical specific_pmids list is 1–20 PMIDs,
        # so the extra NCBI traffic is bounded.
        from .full_text_fetcher import _get_pmcid_for_pmid  # local import to avoid cycle
        oa_mandatory = set()
        paywalled = []
        for pmid in sorted(mandatory_pmids):
            check_cancellation()
            pmcid = _get_pmcid_for_pmid(pmid)
            if pmcid:
                oa_mandatory.add(pmid)
            else:
                paywalled.append(pmid)
        if paywalled:
            papers_excluded_not_oa = len(paywalled)
            emit_log(
                "warn",
                f"OA gate: excluded {len(paywalled)} paywalled PMID(s) from the run",
                f"PMIDs: {', '.join(paywalled[:10])}"
                + ("…" if len(paywalled) > 10 else ""),
            )
        mandatory_pmids = oa_mandatory
    pipeline_stats["papers_excluded_not_oa"] = papers_excluded_not_oa

    # Always include explicitly requested PMIDs (post-gate)
    if mandatory_pmids:
        initial_pmids.update(mandatory_pmids)

    if query:
        check_cancellation()
        if getattr(config, "ENABLE_OA_FILTER", True):
            emit_log(
                "info",
                "Filtering to open-access papers only \u2014 paywalled papers will be excluded",
                "Filter: loattrfull text[sb]",
            )
        # Strategy: get the 100 most relevant (configurable), then later rank those by citations
        relevant_count = getattr(config, "PUBMED_RELEVANT_COUNT", 100)
        query_results = pubmed_data_collector.search_pubmed(query, relevant_count)
        initial_pmids.update(query_results)

    emit_log("info", f"Found {len(initial_pmids)} papers")
    report_progress("Fetching paper details", 20, {"papers_found": len(initial_pmids)})
    check_cancellation()
    emit_log("debug", f"Fetching details for {len(initial_pmids)} PMIDs from PubMed")

    # ── Tracer: record the user selection handoff (pipeline-wide, always recorded when enabled)
    if pipeline_tracer.is_enabled():
        pipeline_tracer.capture(
            "user_selection",
            inputs={
                "query": query or "",
                "specific_pmids": list(specific_pmids or []),
                "specific_authors": list(specific_authors or []),
                "top_n_cited": top_n_cited,
            },
            outputs={
                "deduplicated_pmids_count": len(initial_pmids),
                "mandatory_pmids_count": len(mandatory_pmids),
                "target_pmid_in_selection": pipeline_tracer.matches_any(initial_pmids),
            },
        )

    fetch_start = time.time()
    with pipeline_tracer.stage("pubmed_metadata"):
        paper_details = pubmed_data_collector.fetch_paper_details(list(initial_pmids))
    fetch_duration_ms = (time.time() - fetch_start) * 1000.0
    # ── Tracer: record PubMed metadata fetch for the target PMID
    if pipeline_tracer.is_enabled():
        target = pipeline_tracer.target_pmid()
        entry = paper_details.get(target) if target else None
        pipeline_tracer.capture(
            "pubmed_metadata",
            pmid=target,
            inputs={"pmids_requested": len(initial_pmids)},
            outputs={
                "target_metadata": pipeline_tracer.summarise(entry) if entry else None,
                "retrieved_count": len(paper_details),
            },
            duration_ms=fetch_duration_ms,
        )
    incomplete_count = sum(
        1
        for v in paper_details.values()
        if (v.get("_metadata_completeness", 1) < 1 or v.get("_metadata_warnings"))
    )
    if incomplete_count:
        emit_log(
            "warn",
            f"Metadata quality warning: {incomplete_count}/{len(paper_details)} papers have incomplete metadata",
            "Check 'Metadata Completeness' and 'Metadata Warnings' columns in output",
        )

    # Step 3: Fetch Full Text first for relevance-selected PMIDs
    report_progress("Fetching full text", 30)
    emit_log("info", f"Fetching full text for {len(paper_details)} papers")
    logging.info("STEP 3: Fetching full text for relevance-selected PMIDs...")
    check_cancellation()
    pmids_to_fetch = list(paper_details.keys())
    content_dict_path = _create_unique_filepath("content_dict", "pkl.gz")
    fetch_start = time.time()
    with pipeline_tracer.stage("full_text_fetch"):
        full_text_fetcher.run_fetching(pmids_to_fetch, content_dict_path)
    fetch_duration_ms = (time.time() - fetch_start) * 1000.0

    report_progress("Processing fetched content", 45)
    try:
        with gzip.open(content_dict_path, "rb") as f:
            content_dict = pickle.load(f)
    except Exception:
        logging.warning("Content dictionary is empty. No papers could be fetched.")
        return None

    # Keep only PMIDs successfully scraped
    scraped_pmids = list(content_dict.keys())
    emit_log(
        "info", f"Retrieved full text for {len(scraped_pmids)} of {len(pmids_to_fetch)} papers"
    )

    # ── Tracer: full-text fetch + text cleaning (cleaning already happened inside the fetcher)
    if pipeline_tracer.is_enabled():
        target = pipeline_tracer.target_pmid()
        entry = content_dict.get(target) if target else None
        pipeline_tracer.capture(
            "full_text_fetch",
            pmid=target,
            inputs={"pmids_requested": len(pmids_to_fetch)},
            outputs={
                "scraped_count": len(scraped_pmids),
                "target_present": entry is not None,
                "extraction_method": (entry or {}).get("extraction_method"),
                "content_length": (entry or {}).get("content_length"),
                "quality_score": (entry or {}).get("quality_score"),
                "figures_count": len((entry or {}).get("figures") or []),
                "tables_count": len((entry or {}).get("tables") or []),
                "content_preview": pipeline_tracer.summarise((entry or {}).get("content", "")),
            },
            duration_ms=fetch_duration_ms,
        )
        # The cleaning step happens inside full_text_fetcher; captured as a derived observation.
        pipeline_tracer.capture(
            "text_cleaning",
            pmid=target,
            inputs={"note": "Greek transliteration + non-ASCII strip already applied"},
            outputs={
                "final_content_length": (entry or {}).get("content_length"),
                "ascii_only": True,
            },
        )
    if getattr(config, "ENABLE_FIGURE_ANALYSIS", True):
        papers_with_figures = 0
        total_figures = 0
        for payload in content_dict.values():
            figs = payload.get("figures", []) if isinstance(payload, dict) else []
            if figs:
                papers_with_figures += 1
                total_figures += len(figs)
        if total_figures:
            emit_log("info", f"Discovered {total_figures} figures in {papers_with_figures} papers")

    # Count tables extracted from content_dict
    total_tables = sum(
        len(v.get("tables", []))
        for v in content_dict.values()
        if isinstance(v, dict)
    )
    pipeline_stats["tables_extracted"] = total_tables
    if total_tables:
        papers_with_tables = sum(
            1
            for v in content_dict.values()
            if isinstance(v, dict) and v.get("tables")
        )
        emit_log("info", f"Extracted {total_tables} tables from {papers_with_tables} papers")

    # Track fetch failures
    pipeline_stats["papers_fetch_failed"] = len(pmids_to_fetch) - len(scraped_pmids)

    # Forensic fetch report
    if getattr(config, "FORENSIC_INCLUDE_FETCH_OUTCOMES", True) and content_dict:
        try:
            from .full_text_fetcher import generate_fetch_report

            fetch_report.extend(generate_fetch_report(content_dict))
        except Exception as e:
            logging.warning(f"Forensic fetch report generation failed: {e}")

    # HYBRID PIPELINE: PubTator extraction for high-precision NER
    pubtator_results = {}  # pmid -> HybridExtractionResult
    if getattr(config, "ENABLE_PUBTATOR_EXTRACTION", True) and scraped_pmids:
        report_progress("Running PubTator extraction", 47)
        emit_log("info", "Running PubTator gene extraction")
        logging.info("STEP 3.5: Running PubTator extraction for high-precision gene discovery...")
        check_cancellation()

        try:
            pubtator_tool = PubTatorTool()
            pt_start = time.time()
            with pipeline_tracer.stage("pubtator_ner"):
                batch_results = pubtator_tool.extract_from_pmids(scraped_pmids)
            pt_duration_ms = (time.time() - pt_start) * 1000.0

            for pmid, (genes, variants) in batch_results.items():
                result = HybridExtractionResult(pmid)
                result.pubtator_genes = genes
                result.pubtator_variants = variants
                pubtator_results[pmid] = result

            # ── Tracer: PubTator NER for the target PMID
            if pipeline_tracer.is_enabled():
                target = pipeline_tracer.target_pmid()
                target_result = pubtator_results.get(target) if target else None
                pipeline_tracer.capture(
                    "pubtator_ner",
                    pmid=target,
                    inputs={"scraped_pmids_count": len(scraped_pmids)},
                    outputs={
                        "target_present": target_result is not None,
                        "genes": [g.symbol for g in (target_result.pubtator_genes if target_result else [])],
                        "variants": pipeline_tracer.summarise(
                            [{"text": v.text, "type": v.variant_type, "rsid": v.rsid, "hgvs": v.hgvs}
                             for v in (target_result.pubtator_variants if target_result else [])]
                        ),
                    },
                    meta={"papers_returned": len(pubtator_results)},
                    duration_ms=pt_duration_ms,
                )

            total_pt_genes = sum(len(r.pubtator_genes) for r in pubtator_results.values())
            pt_skipped = len(scraped_pmids) - len(pubtator_results)
            pipeline_stats["pubtator_pmids_skipped"] = pt_skipped
            emit_log(
                "info",
                f"PubTator found {total_pt_genes} genes across {len(pubtator_results)} papers",
                f"{pt_skipped} PMID(s) not returned by PubTator (not indexed or parse error)" if pt_skipped else None,
            )
            logging.info(
                f"PubTator: Found {total_pt_genes} genes across {len(pubtator_results)} papers"
            )
        except Exception as e:
            emit_log("warn", "PubTator extraction failed, continuing without it")
            logging.warning(f"PubTator extraction failed, continuing without it: {e}")

    if not scraped_pmids:
        emit_log("warn", "No full text available, falling back to metadata-only results")
        logging.warning(
            "No scraped papers available after full-text step. Falling back to minimal rows from PubMed metadata."
        )
        # Build minimal rows for initial PMIDs using PubMed details and citation counts
        try:
            # Fetch citations for all initial candidates
            all_candidate_pmids = list(paper_details.keys())
            citation_records = pubmed_data_collector.fetch_citation_counts_with_fallback(
                all_candidate_pmids
            )
            citations_dict = {pmid: rec.get("count", 0) for pmid, rec in citation_records.items()}

            # Rank by citations desc and take top_n_cited
            ranked = sorted(
                all_candidate_pmids, key=lambda p: citations_dict.get(p, 0), reverse=True
            )
            selected = ranked[:top_n_cited] if top_n_cited else ranked

            minimal_rows_only = []
            for pmid in selected:
                base_info = paper_details.get(pmid, {})
                minimal_rows_only.append(build_minimal_row(pmid, base_info, citation_records))

            all_results_df = pd.DataFrame(minimal_rows_only)

            # Save immediately (no user columns beyond core ones yet)
            output_path = _create_unique_filepath("final_enriched_results", "csv")
            all_results_df.to_csv(output_path, index=False, encoding="utf-8")
            debug_path = write_drop_debug_artifact(
                status="metadata_only_no_full_text", output_csv_path=output_path
            )
            report_progress("Completed", 100)
            logging.info(
                f"--- PIPELINE FINISHED WITH MINIMAL ROWS IN {time.time() - start_time:.2f} SECONDS ---"
            )
            return {"local_path": output_path, "debug_path": debug_path}
        except Exception as e:
            logging.error(f"Fallback to minimal rows failed: {e}")
            return None

    # Step 4: Fetch Citations only for scraped PMIDs and select top for AI
    report_progress("Fetching citation counts", 50)
    logging.info("STEP 4: Fetching citations for scraped PMIDs and selecting top papers...")
    check_cancellation()
    cite_start = time.time()
    with pipeline_tracer.stage("citation_fetch"):
        citation_records = pubmed_data_collector.fetch_citation_counts_with_fallback(scraped_pmids)
    citations_dict = {pmid: rec.get("count", 0) for pmid, rec in citation_records.items()}

    # ── Tracer: citation fetch for target PMID
    if pipeline_tracer.is_enabled():
        target = pipeline_tracer.target_pmid()
        rec = citation_records.get(target) if target else None
        pipeline_tracer.capture(
            "citation_fetch",
            pmid=target,
            inputs={"scraped_pmids_count": len(scraped_pmids)},
            outputs={
                "target_record": rec,
                "target_count": (rec or {}).get("count"),
                "source": (rec or {}).get("source"),
            },
            duration_ms=(time.time() - cite_start) * 1000.0,
        )

    scraped_details = {pmid: info for pmid, info in paper_details.items() if pmid in scraped_pmids}
    all_papers_df = pd.DataFrame.from_dict(scraped_details, orient="index")
    if "citations" not in all_papers_df.columns:
        all_papers_df["citations"] = 0
    for pmid, count in citations_dict.items():
        if pmid in all_papers_df.index:
            all_papers_df.at[pmid, "citations"] = count

    report_progress("Selecting top papers", 60)
    # Build processing order: mandatory (if scraped) keep first (original order), then remaining by citations desc
    scraped_set = set(all_papers_df.index)
    ordered_mandatory = [p for p in list(mandatory_pmids) if p in scraped_set]
    remaining_df = (
        all_papers_df.drop(ordered_mandatory, errors="ignore")
        if ordered_mandatory
        else all_papers_df
    )
    ranked_remaining = remaining_df.sort_values(by="citations", ascending=False).index.tolist()
    pmids_to_process = ordered_mandatory + ranked_remaining
    logging.info(
        f"Prepared {len(pmids_to_process)} scraped PMIDs for AI analysis (mandatory first, then ranked by citations)."
    )

    # Abstract screening — forensic logging only (filtering moved to UI preview).
    # The UI now scores papers before the user selects them, so all user-selected
    # papers proceed to AI analysis without silent filtering.
    if getattr(config, "ENABLE_ABSTRACT_SCREENING", True):
        report_progress("Scoring abstracts", 65)
        logging.info(
            "STEP 4.5: Scoring paper abstracts for forensic logging (no filtering — screening moved to UI)..."
        )
        check_cancellation()

        threshold = getattr(config, "ABSTRACT_SCREENING_THRESHOLD", 5)
        would_pass = 0
        would_reject = 0

        for pmid in pmids_to_process:
            base_info = paper_details.get(pmid, {})
            title = base_info.get("title", "")
            abstract = base_info.get("abstract", "")
            should_process, confidence, details = has_genetic_content(abstract, title, threshold)
            if should_process or pmid in mandatory_pmids:
                would_pass += 1
            else:
                would_reject += 1

        logging.info(
            f"Abstract scoring: {would_pass}/{len(pmids_to_process)} would pass threshold "
            f"(all {len(pmids_to_process)} papers proceeding — screening moved to UI)"
        )
        # All papers proceed — no filtering
        pipeline_stats["papers_screened_passed"] = len(pmids_to_process)
        report_progress("Scoring abstracts", 68, {"papers_screened": len(pmids_to_process)})

        # Forensic screening data collection (still useful for debug artifacts)
        if getattr(config, "FORENSIC_INCLUDE_SCREENING", True):
            try:
                from .abstract_screener import screen_papers_with_decisions, decisions_to_dicts

                screening_subset = {
                    pmid: paper_details.get(pmid, {}) for pmid in pmids_to_process
                }
                _, screening_decisions = screen_papers_with_decisions(
                    screening_subset,
                    threshold=threshold,
                    mandatory_pmids=mandatory_pmids,
                )
                forensic_screening.extend(decisions_to_dicts(screening_decisions))
            except Exception as e:
                logging.warning(f"Forensic screening data collection failed: {e}")

        if not pmids_to_process:
            logging.warning(
                "No papers to analyse. Falling back to minimal rows."
            )
            # Build minimal rows for all scraped PMIDs
            minimal_rows_only = []
            for pmid in scraped_pmids[:top_n_cited]:
                base_info = paper_details.get(pmid, {})
                minimal_rows_only.append({**build_minimal_row(pmid, base_info, citation_records)})
            all_results_df = pd.DataFrame(minimal_rows_only)
            output_path = _create_unique_filepath("final_enriched_results", "csv")
            all_results_df.to_csv(output_path, index=False, encoding="utf-8")
            debug_path = write_drop_debug_artifact(
                status="metadata_only_all_rejected_by_screening", output_csv_path=output_path
            )
            report_progress("Completed", 100)
            logging.info(
                f"--- PIPELINE FINISHED WITH MINIMAL ROWS IN {time.time() - start_time:.2f} SECONDS ---"
            )
            return {"local_path": output_path, "debug_path": debug_path}

    report_progress("Analyzing papers with AI", 70)
    emit_log("info", f"Analyzing {len(pmids_to_process)} papers with AI")
    # Step 5: Process each paper using the per-paper extraction coordinator.
    logging.info("STEP 5: Analyzing papers with per-paper extraction pipeline...")
    all_results_df = pd.DataFrame()
    # Sanitize user columns to avoid collisions with core fields
    column_descriptions = _sanitize_user_columns(user_columns)

    total_papers = len(pmids_to_process)
    collected_rows = []
    full_rows_pmids = set()
    minimal_rows = []
    analyzed_attempts = 0

    # Pre-warm persistent worker pool — processes stay alive between papers so the
    # Python interpreter + heavy imports (google-genai, grpc, pandas) are paid once,
    # not once per paper. pool_size=2 means a timed-out worker doesn't block the next paper.
    pool_size = max(1, min(int(getattr(config, "AI_WORKER_POOL_SIZE", 2)), 4))
    worker_pool = mp.Pool(processes=pool_size)
    logging.info(f"AI worker pool created: {pool_size} processes")
    parallel_mode = bool(getattr(config, "PARALLEL_ANALYSIS", False))
    ordered_results = None
    in_flight = {}

    try:
        if parallel_mode:
            emit_log(
                "info",
                f"Parallel AI analysis enabled using the existing worker pool ({pool_size} workers)",
            )
            ordered_results = [None] * total_papers
            submit_idx = 0
            completed_count = 0

            def report_parallel_progress():
                ai_progress = 70 + int((completed_count / max(total_papers, 1)) * 25)
                report_progress(
                    "Analyzing papers with AI",
                    ai_progress,
                    {"papers_analyzed": completed_count},
                )

            def submit_next():
                nonlocal submit_idx, completed_count, analyzed_attempts
                while len(in_flight) < pool_size and submit_idx < total_papers:
                    pmid = pmids_to_process[submit_idx]
                    idx = submit_idx
                    submit_idx += 1
                    analyzed_attempts += 1

                    prepared = _prepare_paper_inputs(
                        pmid, content_dict, paper_details, pubtator_results
                    )
                    short_title = (
                        (prepared["title"][:60] + "...")
                        if len(prepared["title"]) > 60
                        else prepared["title"]
                    )

                    if not prepared["paper_text"]:
                        ordered_results[idx] = {
                            "kind": "minimal",
                            "pmid": pmid,
                            "base_info": prepared["base_info"],
                            "debug": {
                                "pmid": pmid,
                                "status": "no_full_text",
                                "reason": "missing_paper_text_after_fetch",
                                "emitted_rows": 0,
                            },
                        }
                        completed_count += 1
                        report_parallel_progress()
                        continue

                    ctx = {
                        **prepared,
                        "submitted_at": time.time(),
                    }
                    ar = worker_pool.apply_async(
                        _run_pipeline_worker,
                        args=(
                            prepared["paper_text"],
                            column_descriptions,
                            prepared["pt_gene_symbols"],
                            prepared["figure_inputs"],
                            prepared["abstract"],
                            prepared["table_inputs"],
                            pmid,
                            prepared["prepared_content"],
                        ),
                    )
                    in_flight[pmid] = {"idx": idx, "async_result": ar, "ctx": ctx}
                    emit_log(
                        "info",
                        f"[parallel] Submitted paper {idx + 1}/{total_papers}: {short_title}",
                        f"PMID {pmid}",
                    )

            submit_next()

            while in_flight:
                check_cancellation()

                newly_done = [
                    pmid
                    for pmid, info in list(in_flight.items())
                    if info["async_result"].ready()
                ]
                timed_out = []
                for pmid, info in list(in_flight.items()):
                    if pmid in newly_done:
                        continue
                    elapsed = time.time() - info["ctx"]["submitted_at"]
                    if elapsed > config.AI_PER_PAPER_TIMEOUT_SECONDS:
                        timed_out.append(pmid)

                if not newly_done and not timed_out:
                    time.sleep(0.2)
                    continue

                for pmid in newly_done:
                    info = in_flight.pop(pmid)
                    idx = info["idx"]
                    ctx = info["ctx"]
                    try:
                        payload = info["async_result"].get(timeout=0)
                    except Exception as e:
                        payload = {"error": str(e)}
                    paper_df, debug_artifact = _finalize_paper_result(
                        payload,
                        pmid,
                        ctx["base_info"],
                        citation_records,
                        ctx["figure_inputs"],
                        pubtator_results,
                        pipeline_stats,
                        emit_log,
                    )
                    ordered_results[idx] = {
                        "kind": "result",
                        "pmid": pmid,
                        "paper_df": paper_df,
                        "debug": debug_artifact,
                    }
                    completed_count += 1
                    report_parallel_progress()
                    short_title = (
                        (ctx["title"][:60] + "...") if len(ctx["title"]) > 60 else ctx["title"]
                    )
                    emit_log(
                        "info",
                        f"[parallel] Completed paper {idx + 1}/{total_papers}: {short_title}",
                        f"PMID {pmid}",
                    )

                if timed_out:
                    for pmid in timed_out:
                        info = in_flight.pop(pmid, None)
                        if info is None:
                            continue
                        ordered_results[info["idx"]] = {
                            "kind": "timeout",
                            "pmid": pmid,
                            "debug": {
                                "pmid": pmid,
                                "status": "timeout",
                                "reason": f"ai_timeout_{config.AI_PER_PAPER_TIMEOUT_SECONDS}s",
                                "emitted_rows": 0,
                            },
                        }
                        completed_count += 1
                        report_parallel_progress()
                        emit_log(
                            "warn",
                            f"[parallel] Timed out PMID {pmid} — skipping (no retry)",
                        )

                    for pmid, info in list(in_flight.items()):
                        if not info["async_result"].ready():
                            continue
                        try:
                            payload = info["async_result"].get(timeout=0)
                        except Exception as e:
                            payload = {"error": str(e)}
                        paper_df, debug_artifact = _finalize_paper_result(
                            payload,
                            pmid,
                            info["ctx"]["base_info"],
                            citation_records,
                            info["ctx"]["figure_inputs"],
                            pubtator_results,
                            pipeline_stats,
                            emit_log,
                        )
                        ordered_results[info["idx"]] = {
                            "kind": "result",
                            "pmid": pmid,
                            "paper_df": paper_df,
                            "debug": debug_artifact,
                        }
                        completed_count += 1
                        report_parallel_progress()
                        del in_flight[pmid]

                    worker_pool.terminate()
                    worker_pool.join()  # mp.Pool.join() takes no timeout kwarg
                    worker_pool = mp.Pool(processes=pool_size)
                    logging.info(f"AI worker pool recreated: {pool_size} processes")

                    for pmid, info in list(in_flight.items()):
                        ctx = info["ctx"]
                        ctx["submitted_at"] = time.time()
                        new_ar = worker_pool.apply_async(
                            _run_pipeline_worker,
                            args=(
                                ctx["paper_text"],
                                column_descriptions,
                                ctx["pt_gene_symbols"],
                                ctx["figure_inputs"],
                                ctx["abstract"],
                                ctx["table_inputs"],
                                pmid,
                                ctx["prepared_content"],
                            ),
                        )
                        in_flight[pmid] = {
                            "idx": info["idx"],
                            "async_result": new_ar,
                            "ctx": ctx,
                        }
                        emit_log(
                            "info",
                            f"[parallel] Re-submitted PMID {pmid} after pool restart",
                        )

                submit_next()
        else:
            for i, pmid in enumerate(tqdm(pmids_to_process, desc="Processing papers")):
                analyzed_attempts += 1
                ai_progress = 70 + int((i / total_papers) * 25)
                report_progress(
                    "Analyzing papers with AI", ai_progress, {"papers_analyzed": analyzed_attempts}
                )

                prepared = _prepare_paper_inputs(
                    pmid, content_dict, paper_details, pubtator_results
                )
                short_title = (
                    (prepared["title"][:60] + "...") if len(prepared["title"]) > 60 else prepared["title"]
                )
                emit_log(
                    "info", f"Analyzing paper {i + 1}/{total_papers}: {short_title}", f"PMID {pmid}"
                )

                if not prepared["paper_text"]:
                    minimal_rows.append(
                        build_minimal_row(pmid, prepared["base_info"], citation_records)
                    )
                    paper_debug_artifacts.append(
                        {
                            "pmid": pmid,
                            "status": "no_full_text",
                            "reason": "missing_paper_text_after_fetch",
                            "emitted_rows": 0,
                        }
                    )
                    continue

                try:
                    ar = worker_pool.apply_async(
                        _run_pipeline_worker,
                        args=(
                            prepared["paper_text"],
                            column_descriptions,
                            prepared["pt_gene_symbols"],
                            prepared["figure_inputs"],
                            prepared["abstract"],
                            prepared["table_inputs"],
                            pmid,
                            prepared["prepared_content"],
                        ),
                    )
                    submitted_at = time.time()
                    while not ar.ready():
                        check_cancellation()
                        elapsed = time.time() - submitted_at
                        if elapsed > config.AI_PER_PAPER_TIMEOUT_SECONDS:
                            raise mp.TimeoutError()
                        time.sleep(0.2)
                    try:
                        payload = ar.get(timeout=0)
                    except mp.TimeoutError:
                        emit_log("warn", f"AI analysis timed out for PMID {pmid}, skipping")
                        logging.warning(
                            f"AI analysis timed out for PMID {pmid} after {config.AI_PER_PAPER_TIMEOUT_SECONDS}s; skipping"
                        )
                        paper_debug_artifacts.append(
                            {
                                "pmid": pmid,
                                "status": "timeout",
                                "reason": f"ai_timeout_{config.AI_PER_PAPER_TIMEOUT_SECONDS}s",
                                "emitted_rows": 0,
                            }
                        )
                        try:
                            worker_pool.terminate()
                            worker_pool.join()  # mp.Pool.join() takes no timeout kwarg
                        except Exception as e:
                            logging.warning(f"Worker pool cleanup after timeout failed: {e}")
                        worker_pool = mp.Pool(processes=pool_size)
                        logging.info(f"AI worker pool recreated: {pool_size} processes")
                        continue
                except Exception as e:
                    emit_log("error", f"AI analysis failed for PMID {pmid}", str(e))
                    logging.error(f"Failed AI analysis for PMID {pmid}: {e}")
                    paper_df = pd.DataFrame()
                    debug_artifact = {
                        "pmid": pmid,
                        "status": "orchestrator_error",
                        "reason": str(e),
                        "candidate_count": None,
                        "candidates": [],
                        "detail_extraction_status": "",
                        "detail_extraction_error": "",
                        "detail_extraction_rows": None,
                        "validation_drops": [],
                        "strict_gate_drops": [],
                        "evidence_gate_drops": [],
                        "final_associations": [],
                        "emitted_rows": 0,
                    }
                else:
                    paper_df, debug_artifact = _finalize_paper_result(
                        payload,
                        pmid,
                        prepared["base_info"],
                        citation_records,
                        prepared["figure_inputs"],
                        pubtator_results,
                        pipeline_stats,
                        emit_log,
                    )

                all_results_df = _accumulate_result(
                    all_results_df,
                    paper_df,
                    pmid,
                    collected_rows,
                    full_rows_pmids,
                    pipeline_stats,
                    emit_log,
                )
                paper_debug_artifacts.append(debug_artifact)

    except JobCancelledException:
        logging.warning("Job cancellation detected! Stopping new paper processing.")

        # Mark remaining PMIDs as Cancelled
        recorded_pmids = set()
        if ordered_results is not None:
            recorded_pmids = {
                slot["pmid"]
                for slot in ordered_results
                if isinstance(slot, dict) and slot.get("pmid")
            }
        processed_set = (
            set(collected_rows)
            | {r["PMID"] for r in minimal_rows}
            | recorded_pmids
        )
        remaining = [p for p in pmids_to_process if p not in processed_set]

        for pmid in remaining:
            base_info = paper_details.get(pmid, {})
            minimal_rows.append(
                {
                    **build_minimal_row(
                        pmid,
                        base_info,
                        citation_records,
                        gene_group="CANCELLED",
                        variant_name="CANCELLED",
                    ),
                    "Abstract": base_info.get("abstract", "No abstract available - Job Cancelled"),
                }
            )
    finally:
        # Always clean up the worker pool (handles normal completion, cancellation, and errors)
        try:
            worker_pool.terminate()
            worker_pool.join()  # mp.Pool.join() takes no timeout kwarg
        except Exception as e:
            logging.warning(f"Worker pool final cleanup failed: {e}")
        logging.info("AI worker pool terminated")

    if parallel_mode and ordered_results is not None:
        for slot in ordered_results:
            if not isinstance(slot, dict):
                continue
            kind = slot.get("kind")
            if kind == "minimal":
                minimal_rows.append(
                    build_minimal_row(slot["pmid"], slot["base_info"], citation_records)
                )
                paper_debug_artifacts.append(slot["debug"])
            elif kind == "timeout":
                paper_debug_artifacts.append(slot["debug"])
            elif kind == "result":
                all_results_df = _accumulate_result(
                    all_results_df,
                    slot["paper_df"],
                    slot["pmid"],
                    collected_rows,
                    full_rows_pmids,
                    pipeline_stats,
                    emit_log,
                )
                paper_debug_artifacts.append(slot["debug"])

    if all_results_df.empty:
        if minimal_rows:
            logging.warning("AI produced no rows; falling back to minimal rows only.")
            all_results_df = pd.DataFrame(minimal_rows)
        else:
            logging.warning("No validated genes were extracted; saving metadata-only rows.")
            emit_log(
                "warn",
                "No validated genes were extracted",
                "The paper was analyzed, but no gene candidates passed validation and evidence gates.",
            )
            recorded_pmids = {
                str(artifact.get("pmid"))
                for artifact in paper_debug_artifacts
                if artifact.get("pmid")
            } or set(pmids_to_process)
            for pmid in pmids_to_process:
                if pmid not in recorded_pmids:
                    continue
                base_info = paper_details.get(pmid, {})
                minimal_rows.append(
                    {
                        **build_minimal_row(pmid, base_info, citation_records),
                        "extraction_mode": "no_validated_genes",
                        "detail_extraction_error": (
                            "No validated genes were extracted from this paper."
                        ),
                    }
                )
            all_results_df = pd.DataFrame(minimal_rows)

    # HYBRID PIPELINE: NCBI Gene enrichment
    if getattr(config, "ENABLE_NCBI_ENRICHMENT", True) and not all_results_df.empty:
        report_progress("Enriching gene metadata", 92)
        logging.info("STEP 5.5: Enriching genes with NCBI Gene metadata...")

        try:
            # Collect all unique genes that need enrichment
            genes_to_enrich = set()
            if "Gene/Group" in all_results_df.columns:
                genes_to_enrich = set(all_results_df["Gene/Group"].dropna().unique())
                genes_to_enrich = {g for g in genes_to_enrich if g}

            if genes_to_enrich:
                ncbi_tool = NCBIGeneTool()
                symbol_gene_ids = {}
                if "NCBI Gene ID" in all_results_df.columns and "Gene/Group" in all_results_df.columns:
                    for _, row in all_results_df[["Gene/Group", "NCBI Gene ID"]].dropna(how="all").iterrows():
                        symbol = str(row.get("Gene/Group") or "").strip()
                        gene_id = str(row.get("NCBI Gene ID") or "").strip()
                        if symbol and gene_id and ";" not in gene_id:
                            symbol_gene_ids[symbol] = gene_id
                with pipeline_tracer.stage("ncbi_enrichment"):
                    gene_metadata = ncbi_tool.enrich_gene_symbols(
                        list(genes_to_enrich),
                        symbol_gene_ids=symbol_gene_ids,
                    )
                all_results_df = _apply_ncbi_metadata_columns(all_results_df, gene_metadata)

                logging.info(
                    f"NCBI enrichment: Added metadata for {len(gene_metadata)}/{len(genes_to_enrich)} genes"
                )
        except Exception as e:
            logging.warning(f"NCBI enrichment failed, continuing without it: {e}")

    # Enforce desired output policy:
    # - Include up to 'desired' fully reviewed studies (AI-produced rows) first
    # - Append minimal rows for remaining scraped PMIDs at the end (do not count toward N)
    desired = top_n_cited
    if desired:
        # Order full-review PMIDs by processing order
        ordered_full = []
        for pmid in collected_rows:
            if pmid in full_rows_pmids and pmid not in ordered_full:
                ordered_full.append(pmid)
        selected_full_list = ordered_full[:desired]
        selected_full_set = set(selected_full_list)

        # Keep only rows for selected full-review PMIDs (up to N)
        full_df = all_results_df[all_results_df["PMID"].isin(selected_full_set)].reset_index(
            drop=True
        )

        # Build minimal rows for PMIDs that had no full rows (in original processing order)
        minimal_append = []
        selected_full_or_minimal_pmids = set(selected_full_set)
        for mr in minimal_rows:
            if mr["PMID"] in selected_full_or_minimal_pmids:
                continue
            minimal_append.append(mr)
            selected_full_or_minimal_pmids.add(mr["PMID"])

        if minimal_append:
            all_results_df = pd.concat([full_df, pd.DataFrame(minimal_append)], ignore_index=True)
        else:
            all_results_df = full_df

    report_progress("Finalizing results", 95)

    # De-duplicate: If multiple rows within the same PMID and gene have identical user fields, aggregate variant names
    try:
        if not all_results_df.empty:
            core_cols = [
                "PMID",
                "DOI",
                "Gene/Group",
                "Variant Name",
                "Study Title",
                "Authors",
                "Publication Year",
                "Journal Name",
                "Author Affiliations",
                "Citations",
                "Citation Source",
                "Citation Retrieved At",
                "iCite Citations",
                "Semantic Scholar Citations",
                "Figure Count",
                "Figure Analysis Enabled",
                "Candidate Source",
                "Normalization Applied",
                "Validation Outcome",
                "Association Type",
                "Association Group",
                "Dropped By Gate",
            ]
            # Identify user columns (including their citation pairs) that are not core or metadata
            metadata_suffixes = [
                "_citation_valid",
                "_citation_confidence",
                "_citation_details",
                "validation_confidence",
                "validation_source",
                "validation_suggestions",
                "context_flash_fits",
                "context_pro_fits",
                "context_original_tokens",
                "context_modifications",
                "context_truncation_applied",
            ]

            user_cols = [
                c
                for c in all_results_df.columns
                if c not in core_cols
                and not any(c.endswith(s) or c == s for s in metadata_suffixes)
            ]

            # Grouping keys exclude 'Variant Name' so we can aggregate it when other fields are identical
            group_keys = [
                "PMID",
                "Gene/Group",
                "Study Title",
                "Authors",
                "Publication Year",
                "Journal Name",
            ] + user_cols

            # Only dedup rows that actually have a gene (skip minimal rows lacking gene info)
            has_gene_mask = all_results_df["Gene/Group"].fillna("") != ""
            df_full = all_results_df[has_gene_mask].copy()
            df_minimal = all_results_df[~has_gene_mask].copy()

            if not df_full.empty:
                agg_map = {
                    c: "first" for c in df_full.columns if c not in group_keys + ["Variant Name"]
                }

                # Aggregate variant names: unique; sorted; semicolon-joined
                def _agg_variants(series):
                    vals = {str(v) for v in series if str(v).strip()}
                    return "; ".join(sorted(vals))

                agg_map["Variant Name"] = _agg_variants

                df_full = df_full.groupby(group_keys, dropna=False, as_index=False).agg(agg_map)

                # Recombine full and minimal, preserving original order later via sort key
                all_results_df = pd.concat([df_full, df_minimal], ignore_index=True)
    except Exception as e:
        logging.warning(f"Deduplication step skipped due to error: {e}")

    # Ensure unique columns before reordering to prevent reindex errors
    all_results_df = _ensure_unique_columns(all_results_df)

    # Reorder columns for better readability
    core_columns = [
        "Gene/Group",
        "Variant Name",
        "Gene Source",
        "NCBI Gene ID",
        "Gene Full Name",
        "Gene Aliases",
        "Gene Biotype",
        "Chromosome",
        "Candidate Source",
        "Normalization Applied",
        "Validation Outcome",
        "Association Type",
        "Association Group",
        "Dropped By Gate",
        "PMID",
        "DOI",
        "Study Title",
        "Authors",
        "Publication Year",
        "Journal Name",
        "Author Affiliations",
        "Citations",
        "Citation Source",
        "Citation Retrieved At",
        "iCite Citations",
        "Semantic Scholar Citations",
        "Figure Count",
        "Figure Analysis Enabled",
        "Metadata Completeness",
        "Metadata Warnings",
    ]

    # User-defined columns (exclude core and metadata columns)
    metadata_suffixes = [
        "_citation_valid",
        "_citation_confidence",
        "_citation_details",
        "validation_confidence",
        "validation_source",
        "validation_suggestions",
        "context_flash_fits",
        "context_pro_fits",
        "context_original_tokens",
        "context_modifications",
        "context_truncation_applied",
    ]

    user_columns_raw = [
        col
        for col in all_results_df.columns
        if col not in core_columns
        and not any(col.endswith(suffix) or col == suffix for suffix in metadata_suffixes)
    ]
    # Pair user columns with their citation columns if present
    user_columns_list = []
    used = set()
    for col in user_columns_raw:
        if col in used:
            continue
        citation_col = f"{col} Citation"
        user_columns_list.append(col)
        if citation_col in all_results_df.columns:
            user_columns_list.append(citation_col)
            used.add(citation_col)
        used.add(col)

    # Metadata columns at the end
    metadata_columns = [
        col
        for col in all_results_df.columns
        if any(col.endswith(suffix) or col == suffix for suffix in metadata_suffixes)
    ]

    # Build final column order
    final_column_order = []
    for col in core_columns:
        if col in all_results_df.columns:
            final_column_order.append(col)
    final_column_order.extend(user_columns_list)
    final_column_order.extend(metadata_columns)

    all_results_df = all_results_df[final_column_order]

    # Step 6: Save final results — primary CSV + metadata CSV + Excel + JSON
    output_path = _create_unique_filepath("final_enriched_results", "csv")
    with pipeline_tracer.stage("output_writer"):
        primary_path, metadata_path, excel_path, json_path = _write_split_output(
            df=all_results_df,
            output_path=output_path,
            user_cols=user_columns_raw,
        )
        candidate_audit_path = write_candidate_audit_artifact(output_csv_path=primary_path)
        debug_path = write_drop_debug_artifact(status="completed", output_csv_path=primary_path)
        if pipeline_tracer.is_enabled():
            pipeline_tracer.capture(
                "output_writer",
                inputs={"rows": int(len(all_results_df)), "columns": list(all_results_df.columns)},
                outputs={
                    "primary_path": primary_path,
                    "metadata_path": metadata_path,
                    "excel_path": excel_path,
                    "json_path": json_path,
                    "candidate_audit_path": candidate_audit_path,
                    "debug_path": debug_path,
                },
            )

    total_genes = pipeline_stats.get("genes_extracted", 0)
    total_analyzed = analyzed_attempts
    elapsed = time.time() - start_time
    emit_log(
        "info",
        f"Pipeline complete: {total_genes} genes from {total_analyzed} papers ({elapsed:.0f}s)",
    )
    quota_rows = int(pipeline_stats.get("quota_limited_rows", 0) or 0)
    if quota_rows:
        emit_log(
            "warn",
            f"Gemini quota/rate limit affected {quota_rows} output rows",
            "Treat this run as incomplete and rerun after quota resets or with a paid/higher-limit key.",
        )
    report_progress("Completed", 100)
    logging.info(f"--- PIPELINE FINISHED IN {elapsed:.2f} SECONDS ---")

    # ── Tracer: merge worker partials and write trace_<pmid>.json
    trace_path = None
    if pipeline_tracer.is_enabled():
        # Uninstall the function tracer before final I/O so the collector's own
        # work doesn't flood the trace with write-path noise.
        pipeline_tracer.uninstall_function_tracer()
        try:
            target = pipeline_tracer.target_pmid()
            out = os.path.join(config.OUTPUT_DIR, f"trace_{target}.json")
            written = pipeline_tracer.collect_and_write(target, out)
            if written:
                trace_path = str(written)
                emit_log("info", f"Pipeline trace written to {trace_path}")
        except Exception as e:
            logging.warning(f"Trace collection failed: {e}")

    result = {
        "local_path": primary_path,
        "metadata_path": metadata_path,
        "excel_path": excel_path,
        "json_path": json_path,
        "candidate_audit_path": candidate_audit_path,
        "debug_path": debug_path,
        # F10b: explicit alias consumed by the Results banner. Mirrors debug_path
        # so the frontend doesn't have to know that "debug" and "drop_debug" refer
        # to the same artifact.
        "drop_debug_path": debug_path,
    }
    if quota_rows:
        result["warning"] = (
            f"Gemini quota/rate limit affected {quota_rows} output rows. "
            "Rows marked REVIEW/skeleton are fallback evidence, not full AI extraction."
        )
    if trace_path:
        result["trace_path"] = trace_path
    return result
